from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .recommend import build_recommendation
from .schema import sheet_by_name
from .search import LocalSemanticIndex


LAB_TABS = (
    "Master Reagents",
    "Experiments",
    "Daily Log",
    "Formulations",
    "Results",
    "Literature Evidence",
    "Agent Suggestions",
    "Daily Reviews",
    "Process Knowledge",
    "Controlled Vocab",
    "Agent Config",
)


def rows_from_values(values: list[list[Any]]) -> list[dict[str, Any]]:
    if not values:
        return []
    headers = [str(header).strip() for header in values[0]]
    rows: list[dict[str, Any]] = []
    for raw_row in values[1:]:
        row = {
            header: normalize_cell(raw_row[index] if index < len(raw_row) else "")
            for index, header in enumerate(headers)
            if header
        }
        if any(value not in ("", None) for value in row.values()):
            rows.append(row)
    return rows


def values_from_rows(headers: list[str], rows: list[dict[str, Any]]) -> list[list[Any]]:
    return [[row.get(header, "") for header in headers] for row in rows]


def normalize_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def load_workbook_tables(path: str | Path) -> dict[str, list[dict[str, Any]]]:
    workbook = load_workbook(Path(path).expanduser(), data_only=True)
    tables: dict[str, list[dict[str, Any]]] = {}
    for tab in LAB_TABS:
        if tab not in workbook.sheetnames:
            tables[tab] = []
            continue
        worksheet = workbook[tab]
        values = [
            [cell for cell in row]
            for row in worksheet.iter_rows(values_only=True)
            if any(cell not in (None, "") for cell in row)
        ]
        tables[tab] = rows_from_values(values)
    return tables


def build_experiment_entry_from_tables(
    tables: dict[str, list[dict[str, Any]]],
    experiment_id: str,
) -> dict[str, Any]:
    experiment = first_matching(tables.get("Experiments", []), "experiment_id", experiment_id)
    if not experiment:
        raise ValueError(f"Experiment {experiment_id!r} was not found in Experiments.")

    reagents_by_id = {
        str(row.get("reagent_id", "")): row
        for row in tables.get("Master Reagents", [])
        if row.get("reagent_id")
    }
    formulation = [
        enrich_formulation_row(row, reagents_by_id)
        for row in matching_rows(tables.get("Formulations", []), "experiment_id", experiment_id)
    ]
    observations = matching_rows(tables.get("Daily Log", []), "experiment_id", experiment_id)
    results = matching_rows(tables.get("Results", []), "experiment_id", experiment_id)
    evidence = linked_literature(
        tables.get("Literature Evidence", []),
        str(experiment.get("linked_literature_ids", "")),
    )

    return {
        "experiment_id": experiment.get("experiment_id", experiment_id),
        "date": experiment.get("date", ""),
        "project": experiment.get("project", ""),
        "process_type": experiment.get("process_type", ""),
        "objective": experiment.get("objective", ""),
        "hypothesis": experiment.get("hypothesis", ""),
        "operator": experiment.get("operator", ""),
        "status": experiment.get("status", ""),
        "planned_next_step": experiment.get("planned_next_step", ""),
        "summary": experiment.get("summary", ""),
        "formulation": formulation,
        "observations": observations,
        "results": results,
        "literature_evidence": evidence,
    }


def enrich_formulation_row(row: dict[str, Any], reagents_by_id: dict[str, dict[str, Any]]) -> dict[str, Any]:
    enriched = dict(row)
    reagent_id = str(row.get("reagent_id", ""))
    reagent = reagents_by_id.get(reagent_id)
    if reagent:
        enriched["reagent"] = reagent
        for key in (
            "name",
            "common_name",
            "category",
            "molecular_weight_g_mol",
            "density_g_mL",
            "purity_fraction",
        ):
            if key in reagent and f"reagent_{key}" not in enriched:
                enriched[f"reagent_{key}"] = reagent[key]
    return enriched


def suggestion_to_row_dict(suggestion: dict[str, Any]) -> dict[str, Any]:
    plan = suggestion.get("proposed_experiment_plan") if isinstance(suggestion, dict) else None
    return {
        "suggestion_id": suggestion.get("suggestion_id", ""),
        "created_at": suggestion.get("created_at", ""),
        "experiment_id": suggestion.get("experiment_id", ""),
        "recommendation_type": suggestion.get("recommendation_type", ""),
        "rationale": suggestion.get("rationale", ""),
        "proposed_change": suggestion.get("proposed_change", ""),
        "expected_effect": suggestion.get("expected_effect", ""),
        "linked_evidence_ids": ",".join(suggestion.get("linked_evidence_ids", [])),
        "proposed_experiment_id": plan.get("suggested_experiment_id", "") if isinstance(plan, dict) else "",
        "proposed_plan_json": json.dumps(plan, sort_keys=True) if isinstance(plan, dict) else "",
        "safety_check": suggestion.get("safety_check", ""),
        "confidence": suggestion.get("confidence", ""),
        "status": suggestion.get("status", "draft"),
    }


def suggestion_to_values(suggestion: dict[str, Any]) -> list[Any]:
    headers = list(sheet_by_name("Agent Suggestions").headers)
    row = suggestion_to_row_dict(suggestion)
    return [row.get(header, "") for header in headers]


def append_suggestion_to_workbook(
    workbook_path: str | Path,
    suggestion: dict[str, Any],
    output_path: str | Path | None = None,
) -> Path:
    source = Path(workbook_path).expanduser().resolve()
    destination = Path(output_path).expanduser().resolve() if output_path else source
    workbook = load_workbook(source)
    worksheet = workbook["Agent Suggestions"]
    headers = [cell.value for cell in worksheet[1]]
    row_dict = suggestion_to_row_dict(suggestion)
    worksheet.append([row_dict.get(str(header), "") for header in headers])
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return destination


def append_rows_to_workbook(
    workbook_path: str | Path,
    sheet_name: str,
    rows: list[dict[str, Any]],
    output_path: str | Path | None = None,
) -> Path:
    source = Path(workbook_path).expanduser().resolve()
    destination = Path(output_path).expanduser().resolve() if output_path else source
    workbook = load_workbook(source)
    worksheet = workbook[sheet_name]
    headers = [str(cell.value) for cell in worksheet[1]]
    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return destination


def update_workbook_rows_by_key(
    workbook_path: str | Path,
    sheet_name: str,
    updates: list[dict[str, Any]],
    output_path: str | Path | None = None,
) -> Path:
    source = Path(workbook_path).expanduser().resolve()
    destination = Path(output_path).expanduser().resolve() if output_path else source
    workbook = load_workbook(source)
    worksheet = workbook[sheet_name]
    headers = [str(cell.value) for cell in worksheet[1]]
    header_index = {header: index + 1 for index, header in enumerate(headers)}
    for update in updates:
        key_field = str(update.get("key_field", ""))
        key_value = str(update.get("key_value", ""))
        target_field = str(update.get("field", ""))
        if target_field not in header_index:
            continue
        row_number = int(update.get("row_number", 0) or 0)
        if row_number >= 2:
            worksheet.cell(row=row_number, column=header_index[target_field], value=update.get("value", ""))
            continue
        if key_field not in header_index:
            continue
        for row_number in range(2, worksheet.max_row + 1):
            if str(worksheet.cell(row=row_number, column=header_index[key_field]).value or "") == key_value:
                worksheet.cell(row=row_number, column=header_index[target_field], value=update.get("value", ""))
                break
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return destination


def suggest_from_workbook(
    workbook_path: str | Path,
    experiment_id: str,
    index: LocalSemanticIndex | None = None,
) -> dict[str, Any]:
    tables = load_workbook_tables(workbook_path)
    entry = build_experiment_entry_from_tables(tables, experiment_id)
    return build_recommendation(entry, index)


def save_entry_from_workbook(
    workbook_path: str | Path,
    experiment_id: str,
    output_path: str | Path,
) -> Path:
    tables = load_workbook_tables(workbook_path)
    entry = build_experiment_entry_from_tables(tables, experiment_id)
    output = Path(output_path).expanduser().resolve()
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(entry, indent=2) + "\n", encoding="utf-8")
    return output


def matching_rows(rows: list[dict[str, Any]], key: str, value: str) -> list[dict[str, Any]]:
    return [row for row in rows if str(row.get(key, "")) == value]


def first_matching(rows: list[dict[str, Any]], key: str, value: str) -> dict[str, Any] | None:
    for row in rows:
        if str(row.get(key, "")) == value:
            return row
    return None


def linked_literature(rows: list[dict[str, Any]], linked_ids: str) -> list[dict[str, Any]]:
    ids = {item.strip() for item in linked_ids.split(",") if item.strip()}
    if not ids:
        return []
    return [row for row in rows if str(row.get("evidence_id", "")) in ids]
