from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .schema import (
    CONTROLLED_VOCAB_VALIDATIONS,
    SHEETS,
)


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBTLE_FILL = PatternFill("solid", fgColor="D9EAF7")


def build_workbook(include_examples: bool = True) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)

    for spec in SHEETS:
        worksheet = workbook.create_sheet(spec.name)
        worksheet.append(list(spec.headers))
        if include_examples:
            for row in spec.example_rows:
                worksheet.append(list(row))

        for cell in worksheet[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            column = spec.columns[cell.column - 1]
            required = " Required." if column.required else ""
            cell.comment = Comment(f"{column.description}{required}", "lab-notebook-agent")

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.sheet_view.showGridLines = True
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value)
            sample_values = [str(cell.value) for cell in column_cells[:20] if cell.value is not None]
            width = min(max([len(header), *(len(value) for value in sample_values)] + [12]) + 2, 45)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    add_validations(workbook)
    add_workflow_note(workbook)
    return workbook


def save_workbook(path: str | Path, include_examples: bool = True) -> Path:
    output = Path(path).expanduser().resolve()
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = build_workbook(include_examples=include_examples)
    workbook.save(output)
    return output


def add_validations(workbook: Workbook) -> None:
    for sheet_name, fields in CONTROLLED_VOCAB_VALIDATIONS.items():
        worksheet = workbook[sheet_name]
        headers = [cell.value for cell in worksheet[1]]
        for field, allowed_values in fields.items():
            column_letter = get_column_letter(headers.index(field) + 1)
            formula = '"' + ",".join(allowed_values) + '"'
            validation = DataValidation(type="list", formula1=formula, allow_blank=True)
            validation.error = "Choose a value from the controlled vocabulary."
            validation.errorTitle = "Invalid value"
            validation.prompt = "Use the controlled vocabulary for this field."
            validation.promptTitle = "Controlled vocabulary"
            worksheet.add_data_validation(validation)
            validation.add(f"{column_letter}2:{column_letter}1000")


def add_workflow_note(workbook: Workbook) -> None:
    worksheet = workbook["Agent Config"]
    worksheet.append(
        (
            "workflow_note",
            (
                "Enter reagents in Master Reagents, one experiment row in "
                "Experiments, formulation rows in Formulations, observations "
                "in Daily Log, and measurements in Results."
            ),
            (
                "Agent Suggestions should be treated as drafts until reviewed "
                "by a human."
            ),
        )
    )
    for cell in worksheet[worksheet.max_row]:
        cell.fill = SUBTLE_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    worksheet["A1"].comment = Comment(
        "Enter reagents in Master Reagents, one experiment row in Experiments, "
        "formulation rows in Formulations, observations in Daily Log, and "
        "measurements in Results. Agent Suggestions should be treated as drafts "
        "until reviewed by a human.",
        "lab-notebook-agent",
    )
