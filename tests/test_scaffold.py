from __future__ import annotations

import json
import subprocess
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from lab_notebook_agent.agent import (
    AgentRunConfig,
    build_agent_report,
    notebook_context_matches,
    next_followup_experiment_id,
    run_workbook_agent,
    select_literature_evidence_for_entry,
    selected_experiment_ids,
)
from lab_notebook_agent.daily_summary import build_daily_summary_report
from lab_notebook_agent.daily_agent import build_daily_agent_run, build_snapshot_daily_agent_run, run_workbook_daily_agent
from lab_notebook_agent.daily_log_results import (
    apply_daily_log_results_report_to_workbook,
    build_daily_log_results_report,
)
from lab_notebook_agent.daily_reviews import daily_review_row_from_run
from lab_notebook_agent.experiment_record import (
    apply_experiment_record_report_to_workbook,
    build_experiment_record_report,
)
from lab_notebook_agent.formulation_normalization import (
    apply_formulation_normalization_report_to_workbook,
    build_formulation_normalization_report,
)
from lab_notebook_agent.google_sheets import (
    audit_report_against_snapshot,
    batch_update_requests_from_report,
    google_setup_audit_from_metadata,
    google_setup_requests_from_metadata,
    sheet_ids_from_snapshot,
    snapshot_capture_plan,
    snapshot_from_tables,
    snapshot_to_tables,
    validate_snapshot,
)
from lab_notebook_agent.google_api import (
    capture_snapshot_from_google_sheets,
    google_api_doctor,
    run_live_google_daily_agent,
    run_live_google_daily_agent_watch,
    run_live_google_daily_log_results_normalization,
    run_live_google_agent,
    run_live_google_experiment_record,
    run_live_google_formulation_normalization,
    run_live_google_material_scaffold,
    run_live_google_plan_materialization,
    run_live_google_recorded_daily_agent,
    run_live_google_setup,
)
from lab_notebook_agent.cli import main, parse_sheet_id_args
from lab_notebook_agent.litscout import (
    evidence_rows_to_values,
    litscout_works_to_evidence_rows,
    load_litscout_export,
    semantic_litscout_work_matches,
)
from lab_notebook_agent.material_scaffold import (
    apply_material_scaffold_report_to_workbook,
    build_material_scaffold_report,
)
from lab_notebook_agent.material_search import build_process_material_search_report
from lab_notebook_agent.materials import audit_experiment_materials, calculate_formulation_row
from lab_notebook_agent.notebook_search import search_notebook_tables
from lab_notebook_agent.planning import (
    apply_plan_materialization_report_to_workbook,
    build_plan_materialization_report,
)
from lab_notebook_agent.preflight import build_experiment_preflight_report
from lab_notebook_agent.prediction import build_litscout_prediction_report, prediction_from_agent_run
from lab_notebook_agent.recommend import build_recommendation, load_entry
from lab_notebook_agent.recorded_daily_agent import (
    build_recorded_daily_agent_run,
    build_snapshot_recorded_daily_agent_run,
    run_workbook_recorded_daily_agent,
)
from lab_notebook_agent.result_analysis import build_result_analysis
from lab_notebook_agent.schema import SHEETS, workbook_contract
from lab_notebook_agent.search import LocalSemanticIndex
from lab_notebook_agent.sheets import (
    append_suggestion_to_workbook,
    build_experiment_entry_from_tables,
    load_workbook_tables,
    rows_from_values,
    suggestion_to_values,
    suggest_from_workbook,
)
from lab_notebook_agent.templates import save_workbook


class ScaffoldTests(unittest.TestCase):
    def test_workbook_template_contains_contract_tabs_and_headers(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            output = save_workbook(Path(tmpdir) / "template.xlsx")
            workbook = load_workbook(output)
            self.assertEqual([sheet.name for sheet in SHEETS], workbook.sheetnames)
            for spec in SHEETS:
                worksheet = workbook[spec.name]
                headers = [cell.value for cell in worksheet[1]]
                self.assertEqual(list(spec.headers), headers)

    def test_workbook_template_validates_controlled_vocab_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            output = save_workbook(Path(tmpdir) / "template.xlsx")
            workbook = load_workbook(output)
            daily_log_validations = validations_by_range(workbook["Daily Log"])
            results_validations = validations_by_range(workbook["Results"])
            daily_review_validations = validations_by_range(workbook["Daily Reviews"])
            self.assertIn("C2:C1000", daily_log_validations)
            self.assertIn("setup", daily_log_validations["C2:C1000"])
            self.assertIn("test", daily_log_validations["C2:C1000"])
            self.assertIn("I2:I1000", results_validations)
            self.assertIn("observed", results_validations["I2:I1000"])
            self.assertIn("planned", results_validations["I2:I1000"])
            self.assertIn("N2:N1000", daily_review_validations)
            self.assertIn("ready_to_apply", daily_review_validations["N2:N1000"])

    def test_daily_log_schema_keeps_existing_columns_as_prefix(self) -> None:
        daily_log = next(spec for spec in SHEETS if spec.name == "Daily Log")
        self.assertEqual(
            [
                "experiment_id",
                "timestamp",
                "process_stage",
                "temperature_C",
                "rpm",
                "pH",
                "solids_percent",
                "particle_size_nm",
                "conversion_percent",
                "viscosity_cP",
                "observation",
                "issue_tags",
                "attachments_url",
            ],
            list(daily_log.headers)[:13],
        )
        self.assertEqual(
            [
                "residual_monomer_percent",
                "polydispersity_index",
                "Tg_C",
                "hold_time_min",
            ],
            list(daily_log.headers)[13:],
        )

    def test_agent_suggestions_schema_keeps_existing_columns_as_prefix(self) -> None:
        agent_suggestions = next(spec for spec in SHEETS if spec.name == "Agent Suggestions")
        self.assertEqual(
            [
                "suggestion_id",
                "created_at",
                "experiment_id",
                "recommendation_type",
                "rationale",
                "proposed_change",
                "expected_effect",
                "linked_evidence_ids",
                "safety_check",
                "confidence",
                "status",
            ],
            list(agent_suggestions.headers)[:11],
        )
        self.assertEqual(
            [
                "proposed_experiment_id",
                "proposed_plan_json",
            ],
            list(agent_suggestions.headers)[11:],
        )

    def test_semantic_search_finds_emulsion_polymerization_material_roles(self) -> None:
        results = LocalSemanticIndex.from_default().search(
            "emulsion polymerization monomers initiator surfactant particle size",
            k=2,
        )
        self.assertTrue(results)
        self.assertIn("emulsion polymerization", results[0].record["process_type"])

    def test_semantic_search_finds_drive_reaction_sheet_patterns(self) -> None:
        results = LocalSemanticIndex.from_default().search(
            "columnar reaction sheet inputs outputs calculations reaction parameters observations",
            k=6,
        )
        result_ids = {result.record["id"] for result in results}
        self.assertIn("workflow.reaction_sheet.columnar_experiments", result_ids)

    def test_semantic_search_finds_google_drive_reaction_families(self) -> None:
        cases = [
            (
                "RAFT target molecular weight AIBN CTA solvent ratio inhibitor",
                "process.raft_polymerization.solution_calculations",
            ),
            (
                "polyurethane NCO index OH fractions hardness Shore A working time viscosity",
                "process.polyurethane_network.stoichiometry",
            ),
            (
                "Nylon 4 pyrrolidone TBuOK TBAHS CO2 NMR formation percent",
                "process.nylon4.pyrrolidone_synthesis",
            ),
            (
                "bio based oil thiolation amination PTSA Dean Stark water production IPDI DBTDL",
                "process.bio_oil_derivatization.amination_thiolation",
            ),
            (
                "RAFT CTA synthesis carbon disulfide thiol potassium hydroxide addition order ice bath",
                "process.cta_synthesis.trithiocarbonate",
            ),
        ]
        for query, expected_id in cases:
            with self.subTest(expected_id=expected_id):
                results = LocalSemanticIndex.from_default().search(query, k=8)
                result_ids = {result.record["id"] for result in results}
                self.assertIn(expected_id, result_ids)

    def test_notebook_search_finds_reagents_and_process_rows(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            result = search_notebook_tables(
                load_workbook_tables(workbook_path),
                "emulsion polymerization surfactant particle size latex stability",
                k=8,
            )
            self.assertGreater(result["summary"]["records_indexed"], 0)
            sheets = {row["sheet"] for row in result["results"]}
            self.assertIn("Master Reagents", sheets)
            self.assertIn("Process Knowledge", sheets)
            self.assertTrue(
                any(row["key_fields"].get("reagent_id") == "S-SDS" for row in result["results"]),
                result["results"],
            )

    def test_process_material_search_groups_emulsion_reagent_candidates(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tables = load_workbook_tables(save_workbook(Path(tmpdir) / "template.xlsx"))
            report = build_process_material_search_report(
                tables,
                process_type="emulsion polymerization",
                experiment_id="EP-001",
            )
            self.assertEqual("lab-notebook-agent-process-material-search.v1", report["schema"])
            self.assertEqual(4, report["summary"]["role_group_count"])
            self.assertEqual([], report["summary"]["required_roles_missing_candidates"])
            roles = {row["role_group"]: row for row in report["roles"]}
            self.assertEqual("M-SKA", roles["monomer"]["candidate_reagents"][0]["reagent_id"])
            self.assertEqual("I-APS", roles["initiator"]["candidate_reagents"][0]["reagent_id"])
            self.assertEqual("S-SDS", roles["surfactant"]["candidate_reagents"][0]["reagent_id"])
            self.assertEqual("present_in_formulation", roles["monomer"]["status"])
            self.assertEqual("optional_no_candidate", roles["aqueous_phase"]["status"])
            self.assertIn("molecular_weight_g_mol", roles["monomer"]["candidate_reagents"][0]["missing_important_fields"])
            self.assertTrue(roles["surfactant"]["process_knowledge_matches"])

    def test_process_material_search_can_include_optional_roles(self) -> None:
        report = build_process_material_search_report(
            {"Master Reagents": [], "Process Knowledge": []},
            process_type="emulsion polymerization",
            include_optional=True,
        )
        roles = {row["role_group"]: row for row in report["roles"]}
        self.assertIn("crosslinker_or_chain_transfer", roles)
        self.assertEqual("optional_no_candidate", roles["crosslinker_or_chain_transfer"]["status"])

    def test_notebook_search_can_filter_to_daily_log_from_snapshot(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tables = load_workbook_tables(save_workbook(Path(tmpdir) / "template.xlsx"))
            snapshot = snapshot_from_tables(tables)
            result = search_notebook_tables(
                snapshot_to_tables(snapshot),
                "coagulum stir shaft particle size high",
                k=3,
                sheets=("Daily Log",),
            )
            self.assertEqual(["Daily Log"], result["searched_sheets"])
            self.assertEqual("Daily Log", result["results"][0]["sheet"])
            self.assertEqual("EP-001", result["results"][0]["key_fields"]["experiment_id"])

    def test_workbook_contract_matches_sheet_specs(self) -> None:
        contract = workbook_contract()
        sheet_names = [sheet["name"] for sheet in contract["sheets"]]
        self.assertEqual([sheet.name for sheet in SHEETS], sheet_names)
        self.assertIn("emulsion polymerization", contract["controlled_vocab"]["process_type"])
        self.assertIn("feed", contract["controlled_vocab"]["process_stage"])
        self.assertIn("observed", contract["controlled_vocab"]["result_quality_flag"])
        self.assertIn("ready_to_apply", contract["controlled_vocab"]["daily_review_status"])

    def test_suggestion_includes_litscout_commands(self) -> None:
        entry = load_entry(Path(__file__).parents[1] / "examples/emulsion_polymerization_entry.json")
        suggestion = build_recommendation(entry)
        self.assertEqual("EP-001", suggestion["experiment_id"])
        self.assertIn("surfactant", suggestion["proposed_change"].lower())
        self.assertEqual(2, len(suggestion["litscout"]["commands"]))
        self.assertIn("litscout search multi", suggestion["litscout"]["commands"][0])
        self.assertIn("material_audit", suggestion)
        self.assertIn("proposed_experiment_plan", suggestion)
        self.assertIn("result_analysis", suggestion)
        self.assertIn("particle_size_high", suggestion["result_analysis"]["signals"])
        self.assertIn("result_support", suggestion["proposed_experiment_plan"])
        self.assertTrue(suggestion["proposed_experiment_plan"]["result_support"]["limiting_metrics"])

    def test_result_analysis_flags_emulsion_outcome_limits(self) -> None:
        entry = load_entry(Path(__file__).parents[1] / "examples/emulsion_polymerization_entry.json")
        entry["results"].append(
            {
                "sample_id": "EP-001-C",
                "measurement_type": "conversion",
                "value": "72",
                "units": "%",
                "quality_flag": "observed",
            }
        )
        analysis = build_result_analysis(entry)
        self.assertEqual("lab-notebook-agent-result-analysis.v1", analysis["schema"])
        self.assertIn("particle_size_high", analysis["signals"])
        self.assertIn("low_conversion", analysis["signals"])
        limiting = {row["metric_key"]: row for row in analysis["limiting_metrics"]}
        self.assertEqual("above_target", limiting["particle_size"]["status"])
        self.assertEqual("below_target", limiting["conversion"]["status"])
        self.assertIn("Outcome limits", analysis["summary"])

    def test_result_analysis_flags_residual_monomer_and_broad_pdi(self) -> None:
        entry = {
            "experiment_id": "EP-POLY",
            "process_type": "emulsion polymerization",
            "observations": [],
            "results": [
                {
                    "sample_id": "EP-POLY-RM",
                    "measurement_type": "residual monomer",
                    "value": "1.8",
                    "units": "%",
                    "quality_flag": "observed",
                },
                {
                    "sample_id": "EP-POLY-PDI",
                    "measurement_type": "polydispersity index",
                    "value": "0.28",
                    "units": "",
                    "quality_flag": "observed",
                },
            ],
        }

        analysis = build_result_analysis(entry)

        self.assertIn("residual_monomer_high", analysis["signals"])
        self.assertIn("low_conversion", analysis["signals"])
        self.assertIn("broad_psd", analysis["signals"])
        limiting = {row["metric_key"]: row for row in analysis["limiting_metrics"]}
        self.assertEqual("above_target", limiting["residual_monomer"]["status"])
        self.assertEqual("above_target", limiting["polydispersity_index"]["status"])
        self.assertIn("process-health", " ".join(analysis["guidance"]))

    def test_result_analysis_uses_structured_daily_log_polymer_outcomes(self) -> None:
        entry = {
            "experiment_id": "EP-DL",
            "process_type": "emulsion polymerization",
            "observations": [
                {
                    "timestamp": "2026-06-09T18:30:00",
                    "process_stage": "test",
                    "residual_monomer_percent": "1.7",
                    "polydispersity_index": "0.27",
                }
            ],
            "results": [],
        }

        analysis = build_result_analysis(entry)

        self.assertIn("residual_monomer_high", analysis["signals"])
        self.assertIn("broad_psd", analysis["signals"])
        limiting = {row["metric_key"]: row for row in analysis["limiting_metrics"]}
        self.assertEqual("above_target", limiting["residual_monomer"]["status"])
        self.assertEqual("above_target", limiting["polydispersity_index"]["status"])

    def test_rows_from_values_drops_blank_rows_and_preserves_headers(self) -> None:
        rows = rows_from_values(
            [
                ["experiment_id", "objective"],
                ["EP-001", "Reduce particle size"],
                ["", ""],
            ]
        )
        self.assertEqual([{"experiment_id": "EP-001", "objective": "Reduce particle size"}], rows)

    def test_workbook_tabs_assemble_experiment_entry(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            tables = load_workbook_tables(workbook_path)
            entry = build_experiment_entry_from_tables(tables, "EP-001")
            self.assertEqual("EP-001", entry["experiment_id"])
            self.assertEqual("emulsion polymerization", entry["process_type"])
            self.assertEqual(3, len(entry["formulation"]))
            self.assertEqual("solketal acrylate", entry["formulation"][0]["reagent_name"])

    def test_workbook_entry_includes_reagent_stock_concentration(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            tables = load_workbook_tables(workbook_path)
            for row in tables["Master Reagents"]:
                if row["reagent_id"] == "I-APS":
                    row["concentration"] = "0.1"
                    row["concentration_units"] = "M"
            for row in tables["Formulations"]:
                if row["reagent_id"] == "I-APS":
                    row["volume_mL"] = "2.5"

            entry = build_experiment_entry_from_tables(tables, "EP-001")

            initiator = next(row for row in entry["formulation"] if row["reagent_id"] == "I-APS")
            self.assertEqual("0.1", initiator["reagent_concentration"])
            self.assertEqual("M", initiator["reagent_concentration_units"])
            calculation = calculate_formulation_row(initiator)
            self.assertEqual(0.25, calculation["derived"]["moles_mmol"])

    def test_experiment_record_report_normalizes_notebook_rows(self) -> None:
        report = build_experiment_record_report(sample_experiment_record())
        self.assertEqual("lab-notebook-agent-experiment-record.v1", report["schema"])
        self.assertEqual("EP-010", report["experiment_id"])
        self.assertEqual(1, report["summary"]["experiment_rows_to_append"])
        self.assertEqual(1, report["summary"]["formulation_rows_to_append"])
        self.assertEqual(2, report["summary"]["daily_log_rows_to_append"])
        self.assertEqual(1, report["summary"]["result_rows_to_append"])
        self.assertEqual("EP-010", report["append_formulations"][0]["experiment_id"])
        self.assertEqual("feed", report["append_daily_log"][0]["process_stage"])
        self.assertEqual("EP-010-R-001", report["append_results"][0]["sample_id"])
        self.assertEqual("observed", report["append_results"][0]["quality_flag"])
        self.assertEqual([], report["warnings"])

    def test_experiment_record_upserts_inline_reagent_metadata(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            record = sample_experiment_record()
            record["formulation"][0]["reagent"] = {
                "reagent_id": "S-SDS",
                "name": "sodium dodecyl sulfate",
                "category": "surfactant",
                "molecular_weight_g_mol": "288.38",
                "density_g_mL": "1.01",
                "supplier": "Sigma",
            }
            report = build_experiment_record_report(record, tables=load_workbook_tables(workbook_path))
            self.assertEqual(0, report["summary"]["master_reagent_rows_to_append"])
            self.assertEqual(2, report["summary"]["master_reagent_cells_to_update"])
            updates = {row["field"]: row["value"] for row in report["update_master_reagents"]}
            self.assertEqual("1.01", updates["density_g_mL"])
            self.assertEqual("Sigma", updates["supplier"])
            self.assertEqual([], report["warnings"])

    def test_experiment_record_warns_on_master_reagent_conflict(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            record = sample_experiment_record()
            record["formulation"][0]["reagent"] = {
                "reagent_id": "S-SDS",
                "molecular_weight_g_mol": "300",
                "density_g_mL": "1.01",
            }
            report = build_experiment_record_report(record, tables=load_workbook_tables(workbook_path))
            self.assertEqual(1, report["summary"]["master_reagent_cells_to_update"])
            self.assertEqual(
                [
                    {
                        "code": "master_reagent_field_conflict",
                        "sheet": "Master Reagents",
                        "reagent_id": "S-SDS",
                        "field": "molecular_weight_g_mol",
                        "existing_value": "288.38",
                        "proposed_value": "300",
                    }
                ],
                report["warnings"],
            )

    def test_experiment_record_appends_new_inline_reagent_metadata(self) -> None:
        record = sample_experiment_record()
        record["formulation"][0]["reagent_id"] = "S-NEW"
        record["formulation"][0]["reagent_name"] = "new nonionic surfactant"
        record["formulation"][0]["reagent_category"] = "surfactant"
        record["formulation"][0]["reagent_molecular_weight_g_mol"] = "650"
        report = build_experiment_record_report(record)
        self.assertEqual(1, report["summary"]["master_reagent_rows_to_append"])
        reagent = report["append_master_reagents"][0]
        self.assertEqual("S-NEW", reagent["reagent_id"])
        self.assertEqual("new nonionic surfactant", reagent["name"])
        self.assertEqual("650", reagent["molecular_weight_g_mol"])

    def test_experiment_record_report_appends_to_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            output_path = Path(tmpdir) / "recorded.xlsx"
            report = build_experiment_record_report(sample_experiment_record())
            apply_experiment_record_report_to_workbook(workbook_path, report, output_workbook=output_path)
            tables = load_workbook_tables(output_path)
            self.assertTrue(any(row["experiment_id"] == "EP-010" for row in tables["Experiments"]))
            self.assertTrue(any(row["experiment_id"] == "EP-010" for row in tables["Daily Log"]))
            self.assertTrue(any(row["sample_id"] == "EP-010-R-001" for row in tables["Results"]))

    def test_experiment_record_apply_updates_existing_master_reagent_cells(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            output_path = Path(tmpdir) / "recorded-reagent-update.xlsx"
            record = sample_experiment_record()
            record["formulation"][0]["reagent"] = {
                "reagent_id": "S-SDS",
                "density_g_mL": "1.01",
                "supplier": "Sigma",
            }
            report = build_experiment_record_report(record, tables=load_workbook_tables(workbook_path))
            apply_experiment_record_report_to_workbook(workbook_path, report, output_workbook=output_path)
            tables = load_workbook_tables(output_path)
            sds = next(row for row in tables["Master Reagents"] if row["reagent_id"] == "S-SDS")
            self.assertEqual("1.01", sds["density_g_mL"])
            self.assertEqual("Sigma", sds["supplier"])

    def test_experiment_record_snapshot_emits_daily_log_google_batch_requests(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            snapshot = snapshot_from_tables(
                load_workbook_tables(workbook_path),
                {
                    "Experiments": 101,
                    "Formulations": 103,
                    "Daily Log": 104,
                    "Results": 102,
                },
            )
            report = build_experiment_record_report(sample_experiment_record())
            audit = audit_report_against_snapshot(report, snapshot)
            self.assertTrue(audit["valid"], audit)
            requests = batch_update_requests_from_report(report, sheet_ids_from_snapshot(snapshot))
            self.assertEqual(
                [101, 103, 104, 102],
                [request["appendCells"]["sheetId"] for request in requests],
            )

    def test_experiment_record_snapshot_emits_master_reagent_update_request(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            snapshot = snapshot_from_tables(
                load_workbook_tables(workbook_path),
                {
                    "Master Reagents": 100,
                    "Experiments": 101,
                    "Formulations": 103,
                    "Daily Log": 104,
                    "Results": 102,
                },
            )
            record = sample_experiment_record()
            record["formulation"][0]["reagent"] = {
                "reagent_id": "S-SDS",
                "density_g_mL": "1.01",
            }
            report = build_experiment_record_report(record, tables=snapshot_to_tables(snapshot))
            audit = audit_report_against_snapshot(report, snapshot)
            self.assertTrue(audit["valid"], audit["errors"])
            requests = batch_update_requests_from_report(report, sheet_ids_from_snapshot(snapshot))
            self.assertEqual(5, len(requests))
            self.assertEqual(100, requests[0]["updateCells"]["start"]["sheetId"])
            self.assertEqual(6, requests[0]["updateCells"]["start"]["columnIndex"])
            self.assertEqual(
                "1.01",
                requests[0]["updateCells"]["rows"][0]["values"][0]["userEnteredValue"]["stringValue"],
            )

    def test_recorded_daily_agent_projects_record_before_suggesting(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tables = load_workbook_tables(save_workbook(Path(tmpdir) / "template.xlsx"))
            run = build_recorded_daily_agent_run(
                tables,
                sample_experiment_record(),
                AgentRunConfig(),
            )
            self.assertEqual("lab-notebook-agent-recorded-daily-run.v1", run["schema"])
            self.assertEqual("EP-010", run["experiment_id"])
            self.assertEqual("2026-06-10", run["review_date"])
            self.assertEqual(["EP-010"], run["selection"]["selected_experiment_ids"])
            self.assertEqual(1, run["summary"]["suggestion_rows_to_append"])
            self.assertEqual(1, run["summary"]["daily_review_rows_to_append"])
            self.assertEqual(2, run["record_report"]["summary"]["daily_log_rows_to_append"])
            self.assertEqual("EP-010", run["daily_agent_run"]["agent_report"]["runs"][0]["experiment_id"])

    def test_recorded_daily_agent_apply_writes_combined_workbook_rows(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            output_path = Path(tmpdir) / "recorded-daily.xlsx"
            run = run_workbook_recorded_daily_agent(
                workbook_path,
                sample_experiment_record(),
                AgentRunConfig(),
                apply=True,
                output_workbook=output_path,
            )
            self.assertTrue(run["applied"])
            tables = load_workbook_tables(output_path)
            ep_010 = next(row for row in tables["Experiments"] if row["experiment_id"] == "EP-010")
            self.assertEqual("needs_review", ep_010["status"])
            self.assertIn("Daily review 2026-06-10", ep_010["summary"])
            self.assertTrue(any(row["experiment_id"] == "EP-010" for row in tables["Formulations"]))
            self.assertEqual(2, sum(1 for row in tables["Daily Log"] if row["experiment_id"] == "EP-010"))
            self.assertEqual(3, sum(1 for row in tables["Results"] if row["experiment_id"] == "EP-010"))
            self.assertEqual(1, sum(1 for row in tables["Agent Suggestions"] if row["experiment_id"] == "EP-010"))
            self.assertTrue(any("EP-010" in str(row["selected_experiment_ids"]) for row in tables["Daily Reviews"]))

    def test_recorded_daily_agent_apply_links_litscout_evidence_for_new_record(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            works_path = write_fake_litscout_export(Path(tmpdir) / "works.json")
            output_path = Path(tmpdir) / "recorded-daily-litscout.xlsx"
            run = run_workbook_recorded_daily_agent(
                workbook_path,
                sample_experiment_record(),
                AgentRunConfig(litscout_export=str(works_path)),
                apply=True,
                output_workbook=output_path,
            )
            agent_run = run["daily_agent_run"]["agent_report"]["runs"][0]
            self.assertEqual("loaded_export", agent_run["litscout_status"]["status"])
            self.assertEqual(
                ["LIT-EP-010-001"],
                agent_run["append_agent_suggestions"][0]["linked_evidence_ids"],
            )
            tables = load_workbook_tables(output_path)
            self.assertEqual(
                1,
                sum(1 for row in tables["Literature Evidence"] if row["evidence_id"] == "LIT-EP-010-001"),
            )
            experiment = next(row for row in tables["Experiments"] if row["experiment_id"] == "EP-010")
            self.assertEqual("LIT-EP-010-001", experiment["linked_literature_ids"])
            suggestion = next(row for row in tables["Agent Suggestions"] if row["experiment_id"] == "EP-010")
            self.assertEqual("LIT-EP-010-001", suggestion["linked_evidence_ids"])

    def test_recorded_daily_agent_snapshot_emits_combined_google_batch(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tables = load_workbook_tables(save_workbook(Path(tmpdir) / "template.xlsx"))
            snapshot = snapshot_from_tables(
                tables,
                {
                    "Experiments": 101,
                    "Formulations": 103,
                    "Daily Log": 104,
                    "Results": 102,
                    "Agent Suggestions": 222,
                    "Daily Reviews": 333,
                },
            )
            run = build_snapshot_recorded_daily_agent_run(
                snapshot,
                sample_experiment_record(),
                AgentRunConfig(),
            )
            self.assertTrue(run["apply_audit"]["valid"], run["apply_audit"])
            self.assertEqual(6, run["summary"]["apply_request_count"])
            self.assertEqual(
                [101, 103, 104, 102, 222, 333],
                [request["appendCells"]["sheetId"] for request in run["batch_update_requests"]],
            )
            appended_experiment = run["apply_report"]["append_experiments"][0]
            self.assertEqual("EP-010", appended_experiment["experiment_id"])
            self.assertEqual("needs_review", appended_experiment["status"])
            self.assertEqual([], run["apply_report"]["update_experiments"])

    def test_recorded_daily_agent_snapshot_batches_litscout_evidence_for_new_record(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tables = load_workbook_tables(save_workbook(Path(tmpdir) / "template.xlsx"))
            works_path = write_fake_litscout_export(Path(tmpdir) / "works.json")
            snapshot = snapshot_from_tables(
                tables,
                {
                    "Experiments": 101,
                    "Formulations": 103,
                    "Daily Log": 104,
                    "Results": 102,
                    "Literature Evidence": 111,
                    "Agent Suggestions": 222,
                    "Daily Reviews": 333,
                },
            )
            run = build_snapshot_recorded_daily_agent_run(
                snapshot,
                sample_experiment_record(),
                AgentRunConfig(litscout_export=str(works_path)),
            )
            self.assertTrue(run["apply_audit"]["valid"], run["apply_audit"])
            self.assertEqual(7, run["summary"]["apply_request_count"])
            self.assertEqual(
                [101, 103, 104, 102, 111, 222, 333],
                [request["appendCells"]["sheetId"] for request in run["batch_update_requests"]],
            )
            agent_run = run["daily_agent_run"]["agent_report"]["runs"][0]
            self.assertEqual("loaded_export", agent_run["litscout_status"]["status"])
            self.assertEqual(
                ["LIT-EP-010-001"],
                agent_run["append_agent_suggestions"][0]["linked_evidence_ids"],
            )
            self.assertEqual("LIT-EP-010-001", agent_run["append_literature_evidence"][0]["evidence_id"])
            self.assertEqual(
                "LIT-EP-010-001",
                run["apply_report"]["append_experiments"][0]["linked_literature_ids"],
            )

    def test_recorded_daily_agent_merges_normalized_formulations_into_appended_rows(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tables = load_workbook_tables(save_workbook(Path(tmpdir) / "template.xlsx"))
            record = sample_experiment_record()
            record["formulation"][0]["reagent"] = {
                "reagent_id": "S-SDS",
                "density_g_mL": "1.4",
                "purity_fraction": "0.5",
            }
            snapshot = snapshot_from_tables(
                tables,
                {
                    "Master Reagents": 100,
                    "Experiments": 101,
                    "Formulations": 103,
                    "Daily Log": 104,
                    "Results": 102,
                    "Agent Suggestions": 222,
                    "Daily Reviews": 333,
                },
            )

            run = build_snapshot_recorded_daily_agent_run(
                snapshot,
                record,
                AgentRunConfig(),
            )

            self.assertTrue(run["apply_audit"]["valid"], run["apply_audit"])
            self.assertEqual(0, run["apply_report"]["summary"]["formulation_cells_to_update"])
            self.assertEqual(2, run["apply_report"]["summary"]["formulation_cells_merged_into_append"])
            self.assertEqual([], run["apply_report"]["update_formulations"])
            formulation = run["apply_report"]["append_formulations"][0]
            self.assertEqual("0.25", formulation["volume_mL"])
            self.assertEqual("0.606838", formulation["moles_mmol"])
            self.assertFalse(
                any(
                    request.get("updateCells", {}).get("start", {}).get("sheetId") == 103
                    for request in run["batch_update_requests"]
                )
            )
            self.assertEqual([], run["apply_report"]["update_experiments"])

    def test_material_audit_detects_emulsion_roles_and_gaps(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            entry = build_experiment_entry_from_tables(load_workbook_tables(workbook_path), "EP-001")
            audit = audit_experiment_materials(entry)
            groups = {group["role_group"]: group for group in audit["role_groups"]}
            self.assertEqual("present", groups["monomer"]["status"])
            self.assertEqual("present", groups["initiator"]["status"])
            self.assertEqual("present", groups["surfactant"]["status"])
            self.assertFalse(audit["ready_for_quantitative_suggestion"])
            self.assertEqual(3, len(audit["quantity_gaps"]))
            self.assertTrue(any(gap["reagent_id"] == "M-SKA" for gap in audit["reagent_property_gaps"]))

    def test_material_audit_accepts_quantified_entry(self) -> None:
        entry = load_entry(Path(__file__).parents[1] / "examples/emulsion_polymerization_entry.json")
        entry["formulation"][0]["mass_g"] = 10
        entry["formulation"][0]["reagent_molecular_weight_g_mol"] = 156.18
        entry["formulation"][0]["reagent_density_g_mL"] = 1.05
        entry["formulation"][1]["mass_g"] = 0.2
        entry["formulation"][1]["reagent_molecular_weight_g_mol"] = 288.38
        entry["formulation"][1]["concentration"] = "active mass"
        entry["formulation"][2]["mass_g"] = 0.1
        entry["formulation"][2]["reagent_molecular_weight_g_mol"] = 228.2
        audit = audit_experiment_materials(entry)
        self.assertTrue(audit["ready_for_quantitative_suggestion"], audit)
        self.assertEqual([], audit["quantity_gaps"])
        monomer_calc = audit["formulation_calculations"][0]
        self.assertAlmostEqual(64.028685, monomer_calc["derived"]["moles_mmol"], places=4)
        self.assertAlmostEqual(9.52381, monomer_calc["derived"]["volume_mL"], places=4)

    def test_experiment_preflight_flags_incomplete_template_material_fields(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            report = build_experiment_preflight_report(
                load_workbook_tables(workbook_path),
                experiment_id="EP-001",
                stage="review",
            )
            self.assertEqual("needs_attention", report["status"])
            self.assertFalse(report["ready_to_run"])
            self.assertFalse(report["ready_for_quantitative_suggestion"])
            checks = {row["name"]: row for row in report["checks"]}
            self.assertEqual("fail", checks["formulation_quantities"]["status"])
            self.assertEqual("fail", checks["reagent_properties"]["status"])
            self.assertEqual("pass", checks["daily_log_observations"]["status"])
            self.assertEqual("pass", checks["results_measurements"]["status"])

    def test_experiment_preflight_passes_completed_review_entry(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            tables = load_workbook_tables(workbook_path)
            complete_template_materials(tables)
            report = build_experiment_preflight_report(tables, experiment_id="EP-001", stage="review")
            self.assertEqual("ready_with_warnings", report["status"])
            self.assertTrue(report["ready_to_run"])
            self.assertTrue(report["ready_for_quantitative_suggestion"])
            self.assertTrue(report["ready_for_agent_suggestion"])
            checks = {row["name"]: row for row in report["checks"]}
            self.assertEqual("warn", checks["literature_evidence"]["status"])
            self.assertEqual("pass", checks["reagent_properties"]["status"])
            self.assertEqual("pass", checks["reagent_safety"]["status"])

    def test_experiment_preflight_requires_reagent_safety_notes_by_default(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            tables = load_workbook_tables(workbook_path)
            complete_template_materials(tables)
            for row in tables["Master Reagents"]:
                if row["reagent_id"] == "M-SKA":
                    row["hazards"] = ""

            report = build_experiment_preflight_report(tables, experiment_id="EP-001", stage="review")

            checks = {row["name"]: row for row in report["checks"]}
            self.assertEqual("fail", checks["reagent_safety"]["status"])
            self.assertFalse(report["ready_to_run"])
            self.assertIn("M-SKA", str(checks["reagent_safety"]["details"]))

    def test_experiment_preflight_can_warn_on_safety_notes_when_configured(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            tables = load_workbook_tables(workbook_path)
            complete_template_materials(tables)
            set_agent_config(tables, "safety_review_required", "false")
            for row in tables["Master Reagents"]:
                if row["reagent_id"] == "M-SKA":
                    row["hazards"] = ""

            report = build_experiment_preflight_report(tables, experiment_id="EP-001", stage="review")

            checks = {row["name"]: row for row in report["checks"]}
            self.assertEqual("warn", checks["reagent_safety"]["status"])
            self.assertTrue(report["ready_to_run"])

    def test_formulation_calculation_derives_mass_and_moles_from_volume(self) -> None:
        calculation = calculate_formulation_row(
            {
                "reagent_id": "M-TEST",
                "target_role": "core_monomer",
                "volume_mL": "2.0",
                "reagent_density_g_mL": "1.05",
                "reagent_molecular_weight_g_mol": "156.18",
            }
        )
        self.assertAlmostEqual(2.1, calculation["derived"]["mass_g"], places=6)
        self.assertAlmostEqual(13.446, calculation["derived"]["moles_mmol"], places=3)
        self.assertEqual([], calculation["missing_for_calculations"])

    def test_formulation_calculation_uses_purity_for_active_moles(self) -> None:
        calculation = calculate_formulation_row(
            {
                "reagent_id": "M-PURITY",
                "target_role": "core_monomer",
                "mass_g": "10",
                "reagent_molecular_weight_g_mol": "100",
                "reagent_density_g_mL": "1.25",
                "reagent_purity_fraction": "0.8",
            }
        )

        self.assertEqual(0.8, calculation["observed"]["purity_fraction"])
        self.assertEqual(8, calculation["derived"]["active_mass_g"])
        self.assertEqual(80, calculation["derived"]["moles_mmol"])
        self.assertEqual(8, calculation["derived"]["volume_mL"])

    def test_formulation_calculation_uses_purity_for_gross_mass_from_moles(self) -> None:
        calculation = calculate_formulation_row(
            {
                "reagent_id": "M-PURITY",
                "target_role": "core_monomer",
                "moles_mmol": "80",
                "reagent_molecular_weight_g_mol": "100",
                "reagent_density_g_mL": "1.25",
                "reagent_purity_fraction": "80%",
            }
        )

        self.assertEqual(0.8, calculation["observed"]["purity_fraction"])
        self.assertEqual(8, calculation["derived"]["active_mass_g"])
        self.assertEqual(10, calculation["derived"]["mass_g"])
        self.assertEqual(8, calculation["derived"]["volume_mL"])

    def test_formulation_calculation_uses_molar_stock_concentration(self) -> None:
        calculation = calculate_formulation_row(
            {
                "reagent_id": "I-APS",
                "target_role": "initiator",
                "volume_mL": "2.5",
                "reagent_concentration": "0.1",
                "reagent_concentration_units": "M",
            }
        )

        self.assertEqual(0.1, calculation["observed"]["concentration"])
        self.assertEqual("M", calculation["observed"]["concentration_units"])
        self.assertEqual(0.25, calculation["derived"]["moles_mmol"])
        self.assertEqual([], calculation["missing_for_calculations"])

    def test_formulation_calculation_uses_mass_stock_concentration_with_mw(self) -> None:
        calculation = calculate_formulation_row(
            {
                "reagent_id": "I-APS",
                "target_role": "initiator",
                "volume_mL": "2.0",
                "reagent_concentration": "10",
                "reagent_concentration_units": "mg/mL",
                "reagent_molecular_weight_g_mol": "288.38",
            }
        )

        self.assertEqual(0.02, calculation["derived"]["active_mass_g"])
        self.assertAlmostEqual(0.069353, calculation["derived"]["moles_mmol"], places=6)
        self.assertNotIn("mass_g", calculation["derived"])
        self.assertEqual([], calculation["missing_for_calculations"])

    def test_formulation_normalization_derives_missing_cells(self) -> None:
        tables = {
            "Master Reagents": [
                {"reagent_id": "M-SKA", "molecular_weight_g_mol": "156.18", "density_g_mL": "1.05"},
            ],
            "Formulations": [
                {
                    "experiment_id": "EP-001",
                    "reagent_id": "M-SKA",
                    "phase": "monomer feed",
                    "target_role": "core_monomer",
                    "mass_g": "10",
                    "volume_mL": "",
                    "moles_mmol": "",
                }
            ],
        }
        report = build_formulation_normalization_report(tables, experiment_ids=("EP-001",))
        self.assertEqual("lab-notebook-agent-formulation-normalization.v1", report["schema"])
        self.assertEqual(2, report["summary"]["formulation_cells_to_update"])
        updates = {row["field"]: row["value"] for row in report["runs"][0]["update_formulations"]}
        self.assertEqual("9.52381", updates["volume_mL"])
        self.assertEqual("64.028685", updates["moles_mmol"])

    def test_formulation_normalization_uses_master_reagent_purity(self) -> None:
        tables = {
            "Master Reagents": [
                {
                    "reagent_id": "M-PURITY",
                    "molecular_weight_g_mol": "100",
                    "density_g_mL": "1.25",
                    "purity_fraction": "0.8",
                },
            ],
            "Formulations": [
                {
                    "experiment_id": "EP-001",
                    "reagent_id": "M-PURITY",
                    "phase": "monomer feed",
                    "target_role": "core_monomer",
                    "mass_g": "10",
                    "volume_mL": "",
                    "moles_mmol": "",
                }
            ],
        }

        report = build_formulation_normalization_report(tables, experiment_ids=("EP-001",))

        updates = {row["field"]: row["value"] for row in report["runs"][0]["update_formulations"]}
        calculation = report["runs"][0]["calculation"]
        self.assertEqual(0.8, calculation["observed"]["purity_fraction"])
        self.assertEqual(8, calculation["derived"]["active_mass_g"])
        self.assertEqual("8", updates["volume_mL"])
        self.assertEqual("80", updates["moles_mmol"])

    def test_formulation_normalization_uses_master_reagent_stock_concentration(self) -> None:
        tables = {
            "Master Reagents": [
                {
                    "reagent_id": "I-APS",
                    "concentration": "0.1",
                    "concentration_units": "M",
                },
            ],
            "Formulations": [
                {
                    "experiment_id": "EP-STOCK",
                    "reagent_id": "I-APS",
                    "phase": "initiator feed",
                    "target_role": "initiator",
                    "volume_mL": "2.5",
                    "moles_mmol": "",
                }
            ],
        }

        report = build_formulation_normalization_report(tables, experiment_ids=("EP-STOCK",))

        updates = {row["field"]: row["value"] for row in report["runs"][0]["update_formulations"]}
        calculation = report["runs"][0]["calculation"]
        self.assertEqual(0.1, calculation["observed"]["concentration"])
        self.assertEqual("M", calculation["observed"]["concentration_units"])
        self.assertEqual(1, report["summary"]["formulation_cells_to_update"])
        self.assertEqual("0.25", updates["moles_mmol"])

    def test_formulation_normalization_derives_wt_percent_from_mass_total(self) -> None:
        tables = {
            "Master Reagents": [
                {"reagent_id": "W-DI", "density_g_mL": "1.0"},
            ],
            "Formulations": [
                {
                    "experiment_id": "EP-001",
                    "reagent_id": "M-SKA",
                    "phase": "monomer feed",
                    "target_role": "core_monomer",
                    "mass_g": "10",
                    "volume_mL": "9.52381",
                    "moles_mmol": "64.028685",
                    "wt_percent": "",
                },
                {
                    "experiment_id": "EP-001",
                    "reagent_id": "I-APS",
                    "phase": "initiator feed",
                    "target_role": "initiator",
                    "mass_g": "0.2",
                    "moles_mmol": "0.693526",
                    "wt_percent": "",
                },
                {
                    "experiment_id": "EP-001",
                    "reagent_id": "W-DI",
                    "phase": "aqueous",
                    "target_role": "solvent",
                    "volume_mL": "89.8",
                    "wt_percent": "",
                },
            ],
        }

        report = build_formulation_normalization_report(tables, experiment_ids=("EP-001",))

        updates = {
            (run["reagent_id"], update["field"]): update["value"]
            for run in report["runs"]
            for update in run["update_formulations"]
        }
        self.assertEqual("10", updates[("M-SKA", "wt_percent")])
        self.assertEqual("0.2", updates[("I-APS", "wt_percent")])
        self.assertEqual("89.8", updates[("W-DI", "mass_g")])
        self.assertEqual("89.8", updates[("W-DI", "wt_percent")])
        water_run = next(run for run in report["runs"] if run["reagent_id"] == "W-DI")
        self.assertEqual(100, water_run["calculation"]["formulation_mass_total_g"])
        self.assertEqual(4, report["summary"]["formulation_cells_to_update"])

    def test_formulation_normalization_apply_writes_updates_to_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            workbook = load_workbook(workbook_path)
            workbook["Master Reagents"]["F2"] = "156.18"
            workbook["Master Reagents"]["G2"] = "1.05"
            workbook["Formulations"]["E2"] = "10"
            workbook.save(workbook_path)
            output_path = Path(tmpdir) / "normalized.xlsx"
            report = build_formulation_normalization_report(
                load_workbook_tables(workbook_path),
                experiment_ids=("EP-001",),
            )
            apply_formulation_normalization_report_to_workbook(workbook_path, report, output_path)
            normalized = load_workbook(output_path)
            self.assertEqual("9.52381", normalized["Formulations"]["F2"].value)
            self.assertEqual("64.028685", normalized["Formulations"]["G2"].value)

    def test_formulation_normalization_snapshot_emits_google_batch_requests(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            workbook = load_workbook(workbook_path)
            workbook["Master Reagents"]["F2"] = "156.18"
            workbook["Master Reagents"]["G2"] = "1.05"
            workbook["Formulations"]["E2"] = "10"
            workbook.save(workbook_path)
            snapshot = snapshot_from_tables(load_workbook_tables(workbook_path), {"Formulations": 103})
            report = build_formulation_normalization_report(snapshot_to_tables(snapshot), experiment_ids=("EP-001",))
            audit = audit_report_against_snapshot(report, snapshot)
            self.assertTrue(audit["valid"], audit["errors"])
            self.assertEqual(2, audit["summary"]["formulation_cells_to_update"])
            requests = batch_update_requests_from_report(report, sheet_ids_from_snapshot(snapshot))
            self.assertEqual(2, len(requests))
            self.assertEqual(103, requests[0]["updateCells"]["start"]["sheetId"])
            self.assertEqual(5, requests[0]["updateCells"]["start"]["columnIndex"])
            self.assertEqual(6, requests[1]["updateCells"]["start"]["columnIndex"])

    def test_suggestion_maps_to_agent_suggestions_row(self) -> None:
        entry = load_entry(Path(__file__).parents[1] / "examples/emulsion_polymerization_entry.json")
        suggestion = build_recommendation(entry)
        values = suggestion_to_values(suggestion)
        headers = list(next(spec for spec in SHEETS if spec.name == "Agent Suggestions").headers)
        self.assertEqual(13, len(values))
        self.assertEqual("EP-001", values[2])
        self.assertEqual("EP-001-FUP-001", values[headers.index("proposed_experiment_id")])
        self.assertIn("surfactant_package", values[headers.index("proposed_plan_json")])
        self.assertIn("formulations", values[headers.index("proposed_plan_json")])
        self.assertEqual("draft", values[headers.index("status")])

    def test_append_suggestion_to_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            entry = load_entry(Path(__file__).parents[1] / "examples/emulsion_polymerization_entry.json")
            suggestion = build_recommendation(entry)
            output_path = append_suggestion_to_workbook(
                workbook_path,
                suggestion,
                Path(tmpdir) / "with-suggestion.xlsx",
            )
            workbook = load_workbook(output_path)
            worksheet = workbook["Agent Suggestions"]
            headers = [cell.value for cell in worksheet[1]]
            self.assertEqual("EP-001", worksheet["C2"].value)
            self.assertEqual("EP-001-FUP-001", worksheet.cell(row=2, column=headers.index("proposed_experiment_id") + 1).value)
            self.assertIn("surfactant_package", worksheet.cell(row=2, column=headers.index("proposed_plan_json") + 1).value)
            self.assertIn("formulations", worksheet.cell(row=2, column=headers.index("proposed_plan_json") + 1).value)
            self.assertEqual("draft", worksheet.cell(row=2, column=headers.index("status") + 1).value)

    def test_suggest_from_workbook_uses_joined_tab_data(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            suggestion = suggest_from_workbook(workbook_path, "EP-001")
        self.assertEqual("EP-001", suggestion["experiment_id"])
        self.assertIn("particle-size", suggestion["expected_effect"])

    def test_material_scaffold_generates_emulsion_starter_placeholders(self) -> None:
        report = build_material_scaffold_report(
            {"Master Reagents": [], "Formulations": []},
            experiment_id="EP-NEW",
            process_type="emulsion polymerization",
        )
        self.assertEqual(4, report["summary"]["role_groups_considered"])
        self.assertEqual(4, report["summary"]["master_reagent_rows_to_append"])
        self.assertEqual(4, report["summary"]["formulation_rows_to_append"])
        roles = {row["target_role"] for row in report["append_formulations"]}
        self.assertIn("core_monomer", roles)
        self.assertIn("initiator", roles)
        self.assertIn("surfactant", roles)
        self.assertIn("solvent", roles)
        monomer_placeholder = next(row for row in report["append_master_reagents"] if row["reagent_id"] == "AUTO-EP-NEW-MONOMER")
        self.assertIn("acrylate monomer", monomer_placeholder["notes"])
        self.assertIn("molecular_weight_g_mol", monomer_placeholder["notes"])
        scaffold_by_role = {row["role_group"]: row for row in report["role_scaffold"]}
        self.assertIn("acrylate monomer", scaffold_by_role["monomer"]["examples"])
        self.assertIn("molecular_weight_g_mol", scaffold_by_role["monomer"]["important_reagent_fields"])

    def test_material_scaffold_reuses_existing_master_reagents(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            report = build_material_scaffold_report(
                load_workbook_tables(workbook_path),
                experiment_id="EP-002",
                process_type="emulsion polymerization",
            )
            self.assertEqual(1, report["summary"]["master_reagent_rows_to_append"])
            self.assertEqual(4, report["summary"]["formulation_rows_to_append"])
            reagent_ids = {row["reagent_id"] for row in report["append_formulations"]}
            self.assertIn("M-SKA", reagent_ids)
            self.assertIn("I-APS", reagent_ids)
            self.assertIn("S-SDS", reagent_ids)

            output_path = Path(tmpdir) / "scaffolded.xlsx"
            apply_material_scaffold_report_to_workbook(workbook_path, report, output_path)
            workbook = load_workbook(output_path)
            self.assertEqual("EP-002", workbook["Formulations"]["A5"].value)

    def test_material_scaffold_uses_ranked_material_candidate(self) -> None:
        tables = {
            "Master Reagents": [
                {
                    "reagent_id": "M-GENERIC",
                    "name": "generic vinyl monomer",
                    "common_name": "generic monomer",
                    "category": "monomer",
                    "role": "monomer",
                },
                {
                    "reagent_id": "M-BA",
                    "name": "butyl acrylate",
                    "common_name": "BA",
                    "category": "monomer",
                    "role": "core_monomer",
                    "molecular_weight_g_mol": "128.17",
                    "density_g_mL": "0.90",
                    "notes": "soft acrylate latex monomer for emulsion polymerization",
                },
            ],
            "Formulations": [],
            "Process Knowledge": [],
        }
        report = build_material_scaffold_report(
            tables,
            experiment_id="EP-RANK",
            process_type="emulsion polymerization",
            query="butyl acrylate latex",
        )
        formulations_by_role = {
            row["target_role"]: row
            for row in report["append_formulations"]
        }
        self.assertEqual("M-BA", formulations_by_role["core_monomer"]["reagent_id"])
        roles = {row["role_group"]: row for row in report["role_scaffold"]}
        self.assertEqual("ranked_process_material_search", roles["monomer"]["selection_method"])
        self.assertEqual("M-BA", roles["monomer"]["selected_candidate"]["reagent_id"])
        self.assertIn("example_token_overlap", roles["monomer"]["selected_candidate"]["match_reasons"])
        self.assertEqual("append_placeholder_master_reagent", roles["initiator"]["action"])

    def test_material_scaffold_uses_process_knowledge_for_placeholder_notes(self) -> None:
        report = build_material_scaffold_report(
            {
                "Master Reagents": [],
                "Formulations": [],
                "Process Knowledge": [
                    {
                        "process_type": "emulsion polymerization",
                        "material_role": "monomer",
                        "typical_examples": "styrene, butyl acrylate",
                        "guidance": "Select monomers around target Tg and latex stability.",
                    }
                ],
            },
            experiment_id="EP-KNOW",
            process_type="emulsion polymerization",
        )

        scaffold_by_role = {row["role_group"]: row for row in report["role_scaffold"]}
        monomer_placeholder = next(
            row for row in report["append_master_reagents"] if row["reagent_id"] == "AUTO-EP-KNOW-MONOMER"
        )
        self.assertTrue(scaffold_by_role["monomer"]["process_knowledge_matches"])
        self.assertIn("styrene", monomer_placeholder["notes"])
        self.assertIn("butyl acrylate", monomer_placeholder["notes"])
        self.assertIn("Select monomers", monomer_placeholder["notes"])

    def test_material_scaffold_snapshot_emits_google_batch_requests(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            snapshot = snapshot_from_tables(
                load_workbook_tables(workbook_path),
                {
                    "Master Reagents": 100,
                    "Formulations": 103,
                    "Literature Evidence": 111,
                    "Agent Suggestions": 222,
                },
            )
            report = build_material_scaffold_report(
                snapshot_to_tables(snapshot),
                experiment_id="EP-002",
                process_type="emulsion polymerization",
            )
            audit = audit_report_against_snapshot(report, snapshot)
            self.assertTrue(audit["valid"], audit["errors"])
            requests = batch_update_requests_from_report(report, sheet_ids_from_snapshot(snapshot))
            self.assertEqual(2, len(requests))
            self.assertEqual(100, requests[0]["appendCells"]["sheetId"])
            self.assertEqual(103, requests[1]["appendCells"]["sheetId"])

    def test_litscout_export_maps_to_literature_evidence_rows(self) -> None:
        works = [
            {
                "title": "Role of anionic and nonionic surfactants on particle size",
                "service": "openalex",
                "author_names": ["A. Author", "B. Author"],
                "year": 2006,
                "doi": "10.1002/app.23717",
                "url": "https://doi.org/10.1002/app.23717",
                "cited_by_count": 10,
                "concepts": [
                    {"display_name": "Emulsion polymerization"},
                    {"display_name": "Particle size"},
                    {"display_name": "Surfactant"},
                ],
            }
        ]
        rows = litscout_works_to_evidence_rows(
            works,
            experiment_id="EP-001",
            query="emulsion polymerization surfactant particle size",
        )
        self.assertEqual("LIT-EP-001-001", rows[0]["evidence_id"])
        self.assertIn("particle_size", rows[0]["relevance_tags"])
        self.assertEqual(11, len(evidence_rows_to_values(rows)[0]))

    def test_litscout_loader_accepts_ndjson_exports(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            export_path = Path(tmpdir) / "works.ndjson"
            export_path.write_text(
                "\n".join(
                    [
                        json.dumps({"title": "Latex feed profile", "service": "openalex"}),
                        json.dumps({"title": "Particle size distribution", "service": "crossref"}),
                    ]
                )
                + "\n",
                encoding="utf-8",
            )

            works = load_litscout_export(export_path)

            self.assertEqual(["Latex feed profile", "Particle size distribution"], [work["title"] for work in works])

    def test_litscout_export_tags_residual_monomer_and_pdi_evidence(self) -> None:
        works = [
            {
                "title": "Residual monomer and PDI control in acrylic emulsion polymerization",
                "service": "openalex",
                "year": 2024,
                "concepts": [
                    {"display_name": "Residual monomer"},
                    {"display_name": "Polydispersity"},
                    {"display_name": "Persulfate initiator"},
                ],
            }
        ]

        rows = litscout_works_to_evidence_rows(
            works,
            experiment_id="EP-001",
            query="residual monomer PDI emulsion polymerization initiator particle size",
        )

        self.assertIn("initiator", rows[0]["relevance_tags"])
        self.assertIn("particle_size", rows[0]["relevance_tags"])

    def test_litscout_ranking_prioritizes_query_specific_polymerization_evidence(self) -> None:
        works = [
            {
                "title": "Non-Ionic Surfactants for Stabilization of Polymeric Nanoparticles for Biomedical Uses",
                "service": "openalex",
                "year": 2021,
                "cited_by_count": 999,
                "concepts": [
                    {"display_name": "Pulmonary surfactant"},
                    {"display_name": "Nanoparticle"},
                    {"display_name": "Emulsion"},
                    {"display_name": "Biocompatibility"},
                ],
            },
            {
                "title": "Coagulative nucleation and particle size distributions in emulsion polymerization",
                "service": "openalex",
                "year": 1984,
                "cited_by_count": 194,
                "concepts": [
                    {"display_name": "Emulsion polymerization"},
                    {"display_name": "Particle size"},
                    {"display_name": "Nucleation"},
                ],
            },
            {
                "title": "Emulsion polymerization: From fundamental mechanisms to process developments",
                "service": "openalex",
                "year": 2004,
                "cited_by_count": 323,
                "keywords": [
                    {"text": "Emulsion polymerization"},
                    {"text": "Polymer chemistry"},
                    {"text": "Latex stability"},
                ],
            },
        ]
        rows = litscout_works_to_evidence_rows(
            works,
            experiment_id="EP-010",
            query="emulsion polymerization surfactant particle size coagulum latex feed",
            limit=2,
        )
        titles = [row["title"] for row in rows]
        self.assertEqual("Coagulative nucleation and particle size distributions in emulsion polymerization", titles[0])
        self.assertIn("Emulsion polymerization: From fundamental mechanisms to process developments", titles)
        self.assertNotIn("Non-Ionic Surfactants for Stabilization of Polymeric Nanoparticles for Biomedical Uses", titles)
        self.assertNotIn("feed", rows[0]["relevance_tags"])
        self.assertIn("stability", rows[0]["relevance_tags"])
        self.assertEqual("high", rows[0]["confidence"])

    def test_litscout_semantic_search_finds_relevant_exported_works(self) -> None:
        works = [
            {
                "title": "Non-Ionic Surfactants for Biomedical Nanoparticle Stabilization",
                "service": "openalex",
                "year": 2022,
                "cited_by_count": 400,
                "concepts": [
                    {"display_name": "Pulmonary surfactant"},
                    {"display_name": "Nanoparticle"},
                    {"display_name": "Biocompatibility"},
                ],
            },
            {
                "title": "Starved-feed latex nucleation in semibatch emulsion polymerization",
                "service": "semantic_scholar",
                "year": 2019,
                "cited_by_count": 24,
                "abstract": "Latex nucleation and monomer feed duration controlled particle distribution.",
                "concepts": [
                    {"display_name": "Emulsion polymerization"},
                    {"display_name": "Latex"},
                    {"display_name": "Nucleation"},
                ],
            },
        ]

        query = "emulsion polymerization latex nucleation monomer feed particle distribution"
        matches = semantic_litscout_work_matches(works, query, k=2)
        self.assertEqual("Starved-feed latex nucleation in semibatch emulsion polymerization", matches[0]["title"])

        rows = litscout_works_to_evidence_rows(works, experiment_id="EP-010", query=query, limit=1)
        self.assertEqual("Starved-feed latex nucleation in semibatch emulsion polymerization", rows[0]["title"])
        self.assertIn("Local semantic rerank score", rows[0]["notes"])

    def test_cli_litscout_semantic_search_reads_export(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            export_path = Path(tmpdir) / "litscout.json"
            output_path = Path(tmpdir) / "matches.json"
            export_path.write_text(
                json_dumps(
                    [
                        {
                            "title": "Biomedical nanoparticle stabilization",
                            "concepts": [{"display_name": "Biocompatibility"}],
                        },
                        {
                            "title": "Latex nucleation during emulsion polymerization feed",
                            "abstract": "Monomer feed profile controlled particle distribution.",
                            "concepts": [{"display_name": "Latex"}],
                        },
                    ]
                ),
                encoding="utf-8",
            )

            exit_code = main(
                [
                    "litscout-semantic-search",
                    "--input",
                    str(export_path),
                    "emulsion polymerization latex monomer feed particle distribution",
                    "-k",
                    "1",
                    "--output",
                    str(output_path),
                ]
            )

            matches = json.loads(output_path.read_text(encoding="utf-8"))
            self.assertEqual(0, exit_code)
            self.assertEqual("Latex nucleation during emulsion polymerization feed", matches[0]["title"])

    def test_literature_evidence_selection_uses_semantic_similarity_for_next_experiment(self) -> None:
        entry = load_entry(Path(__file__).parents[1] / "examples/emulsion_polymerization_entry.json")
        evidence_rows = [
            {
                "evidence_id": "LIT-EP-001-001",
                "title": "Non-Ionic Surfactants for Biomedical Nanoparticle Stabilization",
                "finding": "Surfactants improved unrelated biomedical nanoparticle suspension stability.",
                "relevance_tags": "surfactant",
                "confidence": "high",
            },
            {
                "evidence_id": "LIT-EP-001-002",
                "title": "Coagulative nucleation and particle size distributions in emulsion polymerization",
                "finding": "Latex feed profile reduced coagulum and narrowed particle size distributions.",
                "relevance_tags": "",
                "confidence": "low",
            },
        ]

        selected = select_literature_evidence_for_entry(
            entry,
            evidence_rows,
            query="emulsion polymerization particle size coagulum latex nucleation feed",
            limit=1,
        )

        self.assertEqual("LIT-EP-001-002", selected[0]["evidence_id"])

    def test_recommendation_links_literature_evidence_ids(self) -> None:
        entry = load_entry(Path(__file__).parents[1] / "examples/emulsion_polymerization_entry.json")
        entry["literature_evidence"] = [
            {
                "evidence_id": "LIT-EP-001-001",
                "title": "Surfactant control of emulsion polymer particle size",
                "finding": "Surfactant package can affect particle size and colloidal stability.",
                "relevance_tags": "surfactant,particle_size,stability",
                "confidence": "medium",
            }
        ]
        suggestion = build_recommendation(entry)
        self.assertEqual(["LIT-EP-001-001"], suggestion["linked_evidence_ids"])
        self.assertIn("Literature Evidence rows", suggestion["rationale"])
        self.assertIn("Linked literature tags emphasize", suggestion["rationale"])
        self.assertIn("Literature guidance", suggestion["rationale"])
        self.assertEqual(1, suggestion["literature_context"]["evidence_count"])
        self.assertEqual(1, suggestion["literature_context"]["tag_counts"]["surfactant"])
        self.assertIn("particle_size", suggestion["literature_context"]["relevance_tags"])
        self.assertEqual(["LIT-EP-001-001"], suggestion["proposed_experiment_plan"]["linked_evidence_ids"])
        support = suggestion["proposed_experiment_plan"]["literature_support"]
        self.assertEqual(["LIT-EP-001-001"], support["evidence_ids"])
        self.assertTrue(support["guidance"])

    def test_recommendation_infers_literature_tags_from_findings(self) -> None:
        entry = load_entry(Path(__file__).parents[1] / "examples/emulsion_polymerization_entry.json")
        entry["literature_evidence"] = [
            {
                "evidence_id": "LIT-EP-001-002",
                "title": "Semibatch monomer feed and latex nucleation",
                "finding": "Feed profile changes affected particle size in acrylate latex experiments.",
            }
        ]
        suggestion = build_recommendation(entry)
        context = suggestion["literature_context"]
        self.assertIn("feed", context["relevance_tags"])
        self.assertIn("particle_size", context["relevance_tags"])
        self.assertIn("feed/nucleation", " ".join(context["guidance"]))
        self.assertEqual("LIT-EP-001-002", context["supporting_findings"][0]["evidence_id"])

    def test_structured_emulsion_plan_has_variables_and_capture_fields(self) -> None:
        entry = load_entry(Path(__file__).parents[1] / "examples/emulsion_polymerization_entry.json")
        suggestion = build_recommendation(entry)
        plan = suggestion["proposed_experiment_plan"]
        self.assertEqual("EP-001-FUP-001", plan["suggested_experiment_id"])
        factors = [variable["factor"] for variable in plan["variables"]]
        self.assertIn("surfactant_active_basis_or_feed_duration", factors)
        self.assertIn("surfactant_package", factors)
        self.assertIn("particle_size_nm", plan["measurements"])
        self.assertIn("polydispersity_index", plan["measurements"])
        self.assertIn("residual_monomer_percent", plan["measurements"])
        self.assertTrue(plan["prerequisites"])
        self.assertEqual("planned", plan["sheet_rows"]["experiments"][0]["status"])
        adjustments = plan["planned_formulation_adjustments"]
        self.assertEqual("feed_duration_min", adjustments[0]["field"])
        self.assertEqual("180", str(adjustments[0]["parent_value"]))
        self.assertEqual("225", adjustments[0]["proposed_value"])
        formulation_by_role = {
            row["target_role"]: row
            for row in plan["sheet_rows"]["formulations"]
        }
        self.assertEqual("225", formulation_by_role["core_monomer"]["feed_duration_min"])
        self.assertIn("Applied planned adjustments", formulation_by_role["core_monomer"]["notes"])

    def test_structured_emulsion_plan_adjusts_numeric_surfactant_basis(self) -> None:
        entry = load_entry(Path(__file__).parents[1] / "examples/emulsion_polymerization_entry.json")
        entry["formulation"][1]["mass_g"] = "0.20"
        suggestion = build_recommendation(entry)
        plan = suggestion["proposed_experiment_plan"]
        adjustments = plan["planned_formulation_adjustments"]
        surfactant_adjustment = next(
            adjustment
            for adjustment in adjustments
            if adjustment["target_role"] == "surfactant" and adjustment["field"] == "mass_g"
        )
        self.assertEqual("0.20", surfactant_adjustment["parent_value"])
        self.assertEqual("0.23", surfactant_adjustment["proposed_value"])
        formulation_by_role = {
            row["target_role"]: row
            for row in plan["sheet_rows"]["formulations"]
        }
        self.assertEqual("0.23", formulation_by_role["surfactant"]["mass_g"])
        self.assertEqual(180, formulation_by_role["core_monomer"]["feed_duration_min"])

    def test_structured_emulsion_plan_adjusts_surfactant_stock_volume_basis(self) -> None:
        entry = load_entry(Path(__file__).parents[1] / "examples/emulsion_polymerization_entry.json")
        entry["formulation"][1]["volume_mL"] = "2.5"
        entry["formulation"][1]["concentration"] = "0.1"
        entry["formulation"][1]["concentration_units"] = "M"

        suggestion = build_recommendation(entry)

        plan = suggestion["proposed_experiment_plan"]
        adjustment = next(
            adjustment
            for adjustment in plan["planned_formulation_adjustments"]
            if adjustment["target_role"] == "surfactant" and adjustment["field"] == "volume_mL"
        )
        self.assertEqual("2.5", adjustment["parent_value"])
        self.assertEqual("2.875", adjustment["proposed_value"])
        formulation_by_role = {
            row["target_role"]: row
            for row in plan["sheet_rows"]["formulations"]
        }
        self.assertEqual("2.875", formulation_by_role["surfactant"]["volume_mL"])
        self.assertEqual("0.1", formulation_by_role["surfactant"]["concentration"])
        self.assertEqual("M", formulation_by_role["surfactant"]["concentration_units"])

    def test_structured_emulsion_plan_clears_stale_quantities_after_basis_adjustment(self) -> None:
        entry = load_entry(Path(__file__).parents[1] / "examples/emulsion_polymerization_entry.json")
        entry["formulation"][1]["mass_g"] = "0.20"
        entry["formulation"][1]["volume_mL"] = "0.19"
        entry["formulation"][1]["moles_mmol"] = "0.876"

        suggestion = build_recommendation(entry)

        formulation_by_role = {
            row["target_role"]: row
            for row in suggestion["proposed_experiment_plan"]["sheet_rows"]["formulations"]
        }
        surfactant = formulation_by_role["surfactant"]
        self.assertEqual("0.23", surfactant["mass_g"])
        self.assertEqual("", surfactant["volume_mL"])
        self.assertEqual("", surfactant["moles_mmol"])
        self.assertIn("Cleared dependent quantity fields", surfactant["notes"])

    def test_structured_emulsion_plan_focuses_low_conversion_on_process_health(self) -> None:
        entry = load_entry(Path(__file__).parents[1] / "examples/emulsion_polymerization_entry.json")
        entry["objective"] = "Improve monomer conversion without changing latex chemistry."
        entry["hypothesis"] = "Purge quality or initiator freshness limited conversion."
        entry["observations"] = [
            {
                "timestamp": "2026-06-09T14:35:00",
                "process_stage": "hold",
                "temperature_C": 70,
                "rpm": 250,
                "observation": "Latex remained smooth through the thermal hold.",
            }
        ]
        entry["results"] = [
            {
                "sample_id": "EP-001-L1",
                "measurement_type": "conversion",
                "value": 76,
                "units": "%",
                "quality_flag": "ok",
                "interpretation": "Below target conversion.",
            }
        ]

        suggestion = build_recommendation(entry)
        plan = suggestion["proposed_experiment_plan"]

        self.assertIn("process-health", suggestion["proposed_change"])
        self.assertIn("low conversion", plan["objective"])
        factors = [variable["factor"] for variable in plan["variables"]]
        self.assertIn("initiator_process_health", factors)
        self.assertNotIn("surfactant_package", factors)
        self.assertIn("Conversion moves above", plan["acceptance_criteria"][0])
        planned_experiment = plan["sheet_rows"]["experiments"][0]
        self.assertIn("low conversion", planned_experiment["objective"])
        self.assertNotIn("coagulum", planned_experiment["objective"].lower())

    def test_structured_emulsion_plan_treats_broad_pdi_as_distribution_signal(self) -> None:
        entry = load_entry(Path(__file__).parents[1] / "examples/emulsion_polymerization_entry.json")
        entry["observations"] = [
            {
                "timestamp": "2026-06-09T14:35:00",
                "process_stage": "workup",
                "observation": "Latex was smooth but DLS distribution was broad.",
            }
        ]
        entry["results"] = [
            {
                "sample_id": "EP-001-PDI",
                "measurement_type": "polydispersity index",
                "value": "0.31",
                "units": "",
                "quality_flag": "observed",
            }
        ]
        entry["formulation"][1]["mass_g"] = "0.20"

        suggestion = build_recommendation(entry)

        plan = suggestion["proposed_experiment_plan"]
        self.assertIn("broad_psd", suggestion["result_analysis"]["signals"])
        factors = [variable["factor"] for variable in plan["variables"]]
        self.assertIn("surfactant_active_basis_or_feed_duration", factors)
        surfactant_adjustment = next(
            adjustment
            for adjustment in plan["planned_formulation_adjustments"]
            if adjustment["target_role"] == "surfactant" and adjustment["field"] == "mass_g"
        )
        self.assertEqual("0.23", surfactant_adjustment["proposed_value"])
        self.assertIn("PDI", surfactant_adjustment["rationale"])

    def test_agent_report_appends_evidence_and_linked_suggestion(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            works_path = write_fake_litscout_export(Path(tmpdir) / "works.json")
            tables = load_workbook_tables(workbook_path)
            report = build_agent_report(
                tables,
                AgentRunConfig(experiment_ids=("EP-001",), litscout_export=str(works_path)),
            )
            self.assertEqual("ready", report["runs"][0]["status"])
            self.assertEqual(1, report["summary"]["suggestion_rows_to_append"])
            self.assertEqual(1, report["summary"]["evidence_rows_to_append"])
            self.assertEqual(1, report["summary"]["experiment_cells_to_update"])
            self.assertEqual("linked_literature_ids", report["update_experiments"][0]["field"])
            self.assertEqual("LIT-EP-001-001", report["update_experiments"][0]["value"])
            suggestion = report["runs"][0]["append_agent_suggestions"][0]
            self.assertEqual(["LIT-EP-001-001"], suggestion["linked_evidence_ids"])
            self.assertEqual("loaded_export", report["runs"][0]["litscout_status"]["status"])
            self.assertEqual(1, report["runs"][0]["litscout_status"]["works_count"])
            context_sheets = {row["sheet"] for row in report["runs"][0]["notebook_context_matches"]}
            self.assertIn("Master Reagents", context_sheets)
            self.assertIn("Process Knowledge", context_sheets)

    def test_agent_report_uses_experiment_linked_literature_evidence(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tables = load_workbook_tables(save_workbook(Path(tmpdir) / "template.xlsx"))
            tables["Experiments"][0]["linked_literature_ids"] = "LIT-MANUAL-SURF-001"
            tables["Literature Evidence"].append(
                {
                    "evidence_id": "LIT-MANUAL-SURF-001",
                    "source": "manual",
                    "title": "Reviewed surfactant package note",
                    "finding": "Mixed surfactant packages can improve latex stability and particle size.",
                    "relevance_tags": "surfactant,particle_size,stability",
                    "confidence": "medium",
                }
            )
            report = build_agent_report(tables, AgentRunConfig(experiment_ids=("EP-001",)))
            run = report["runs"][0]
            self.assertEqual("ready", run["status"])
            self.assertEqual("existing_evidence", run["litscout_status"]["status"])
            self.assertEqual([], run["append_literature_evidence"])
            self.assertEqual(0, report["summary"]["evidence_rows_to_append"])
            self.assertEqual(0, report["summary"]["experiment_cells_to_update"])
            suggestion = run["append_agent_suggestions"][0]
            self.assertEqual(["LIT-MANUAL-SURF-001"], suggestion["linked_evidence_ids"])
            self.assertEqual(
                ["LIT-MANUAL-SURF-001"],
                suggestion["proposed_experiment_plan"]["literature_support"]["evidence_ids"],
            )

    def test_agent_report_ranks_existing_literature_evidence_by_relevance_limit(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tables = load_workbook_tables(save_workbook(Path(tmpdir) / "template.xlsx"))
            tables["Experiments"][0]["linked_literature_ids"] = "LIT-OFF-001,LIT-SURF-001,LIT-FEED-001"
            tables["Literature Evidence"].extend(
                [
                    {
                        "evidence_id": "LIT-OFF-001",
                        "source": "manual",
                        "title": "Pulmonary surfactant review",
                        "finding": "Biomedical surfactant systems without polymer latex process guidance.",
                        "relevance_tags": "",
                        "confidence": "high",
                    },
                    {
                        "evidence_id": "LIT-SURF-001",
                        "source": "manual",
                        "title": "Surfactant package control in emulsion polymerization",
                        "finding": "Surfactant package controls latex particle size and coagulum.",
                        "relevance_tags": "surfactant,particle_size,stability",
                        "confidence": "medium",
                    },
                    {
                        "evidence_id": "LIT-FEED-001",
                        "source": "manual",
                        "title": "Semi-batch feed and latex nucleation",
                        "finding": "Monomer feed profile affects particle size in acrylate emulsion polymerization.",
                        "relevance_tags": "feed,particle_size",
                        "confidence": "medium",
                    },
                ]
            )

            report = build_agent_report(
                tables,
                AgentRunConfig(experiment_ids=("EP-001",), evidence_limit=2),
            )

            run = report["runs"][0]
            suggestion = run["append_agent_suggestions"][0]
            self.assertEqual("ready", run["status"])
            self.assertEqual(["LIT-SURF-001", "LIT-FEED-001"], run["selected_literature_evidence_ids"])
            self.assertEqual(["LIT-SURF-001", "LIT-FEED-001"], suggestion["linked_evidence_ids"])
            self.assertEqual(2, suggestion["literature_context"]["evidence_count"])
            self.assertNotIn("LIT-OFF-001", suggestion["linked_evidence_ids"])

    def test_agent_report_links_selected_existing_evidence_when_suggestion_suppressed(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tables = load_workbook_tables(save_workbook(Path(tmpdir) / "template.xlsx"))
            tables["Experiments"][0]["linked_literature_ids"] = ""
            tables["Literature Evidence"].append(
                {
                    "evidence_id": "LIT-EP-001-777",
                    "source": "manual",
                    "title": "Surfactant and particle-size control",
                    "finding": "Surfactant active basis can reduce coagulum and latex particle size.",
                    "relevance_tags": "surfactant,particle_size,stability",
                    "confidence": "high",
                }
            )

            report = build_agent_report(
                tables,
                AgentRunConfig(experiment_ids=("EP-001",), suggestion_confidence_floor="high"),
            )

            run = report["runs"][0]
            self.assertEqual("skipped", run["status"])
            self.assertEqual("suggestion_confidence_below_floor", run["skip_reason"])
            self.assertEqual(["LIT-EP-001-777"], run["selected_literature_evidence_ids"])
            self.assertEqual([], run["append_agent_suggestions"])
            self.assertEqual("linked_literature_ids", report["update_experiments"][0]["field"])
            self.assertEqual("LIT-EP-001-777", report["update_experiments"][0]["value"])

    def test_agent_report_includes_prior_result_history_in_suggestion(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tables = load_workbook_tables(save_workbook(Path(tmpdir) / "template.xlsx"))
            tables["Experiments"].insert(
                0,
                {
                    "experiment_id": "EP-000",
                    "date": "2026-06-01",
                    "project": "SABER CCSP",
                    "process_type": "emulsion polymerization",
                    "objective": "Baseline particle size screen.",
                    "status": "complete",
                    "summary": "Lower particle size but stable latex.",
                },
            )
            tables["Results"].append(
                {
                    "experiment_id": "EP-000",
                    "sample_id": "EP-000-L1",
                    "measurement_type": "DLS particle size",
                    "value": "240",
                    "units": "nm",
                    "quality_flag": "ok",
                    "interpretation": "Within target range.",
                }
            )
            report = build_agent_report(
                tables,
                AgentRunConfig(experiment_ids=("EP-001",)),
            )
            run = report["runs"][0]
            self.assertEqual(1, run["historical_context"]["prior_experiment_count"])
            self.assertIn("EP-000", run["historical_context"]["guidance"][0])
            benchmarks = {row["metric_key"]: row for row in run["historical_context"]["measurement_benchmarks"]}
            self.assertEqual("240", benchmarks["particle_size"]["min"])
            suggestion = run["append_agent_suggestions"][0]
            self.assertIn("Notebook history", suggestion["rationale"])
            self.assertEqual(1, suggestion["historical_context"]["prior_experiment_count"])
            self.assertEqual("EP-000", suggestion["proposed_experiment_plan"]["history_support"]["prior_experiments"][0]["experiment_id"])

    def test_agent_run_cli_history_limit_can_disable_prior_results(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            workbook = load_workbook(workbook_path)
            experiments = workbook["Experiments"]
            experiments.append(
                [
                    "EP-000",
                    "2026-06-01",
                    "SABER CCSP",
                    "emulsion polymerization",
                    "Baseline particle size screen.",
                    "",
                    "",
                    "",
                    "complete",
                    "",
                    "Lower particle size but stable latex.",
                ]
            )
            workbook["Results"].append(
                [
                    "EP-000",
                    "EP-000-L1",
                    "DLS particle size",
                    "intensity average",
                    "240",
                    "nm",
                    "post-feed",
                    "1",
                    "ok",
                    "Within target range.",
                ]
            )
            workbook.save(workbook_path)
            output = Path(tmpdir) / "agent-report.json"
            with patch("builtins.print"):
                exit_code = main(
                    [
                        "agent-run",
                        "--workbook",
                        str(workbook_path),
                        "--experiment-id",
                        "EP-001",
                        "--history-limit",
                        "0",
                        "--report-output",
                        str(output),
                    ]
                )
            report = json.loads(output.read_text(encoding="utf-8"))
            self.assertEqual(0, exit_code)
            self.assertEqual(0, report["runs"][0]["historical_context"]["prior_experiment_count"])
            suggestion = report["runs"][0]["append_agent_suggestions"][0]
            self.assertEqual(0, suggestion["proposed_experiment_plan"]["history_support"]["prior_experiment_count"])

    def test_agent_report_can_run_litscout_and_records_status(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            calls = []

            def fake_run(command: list[str], check: bool) -> None:
                calls.append(command)
                if command[:3] == ["litscout", "sessions", "export"]:
                    output_path = Path(command[command.index("--output") + 1])
                    write_fake_litscout_export(output_path)

            with patch("lab_notebook_agent.agent.subprocess.run", side_effect=fake_run):
                report = build_agent_report(
                    load_workbook_tables(workbook_path),
                    AgentRunConfig(
                        experiment_ids=("EP-001",),
                        run_litscout=True,
                        artifacts_dir=tmpdir,
                    ),
                )

            run = report["runs"][0]
            self.assertEqual("ready", run["status"])
            self.assertEqual("completed", run["litscout_status"]["status"])
            self.assertEqual(1, run["litscout_status"]["works_count"])
            self.assertTrue(run["litscout_export"].endswith("litscout-ep-001.json"))
            self.assertEqual(1, report["summary"]["evidence_rows_to_append"])
            self.assertEqual(2, len(calls))
            self.assertEqual(["litscout", "search", "multi"], calls[0][:3])
            self.assertEqual(run["litscout_query"], calls[0][3])
            self.assertIn("surfactant", calls[0][3])
            self.assertEqual("LIT-EP-001-001", run["litscout_semantic_matches"][0]["evidence_id"])
            self.assertGreater(run["litscout_semantic_matches"][0]["score"], 0)

    def test_agent_report_records_litscout_failure_without_suggestion(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            with patch("lab_notebook_agent.agent.subprocess.run", side_effect=FileNotFoundError()):
                report = build_agent_report(
                    load_workbook_tables(workbook_path),
                    AgentRunConfig(
                        experiment_ids=("EP-001",),
                        run_litscout=True,
                        artifacts_dir=tmpdir,
                    ),
                )

            run = report["runs"][0]
            self.assertEqual("skipped", run["status"])
            self.assertEqual("litscout_failed", run["skip_reason"])
            self.assertEqual("failed", run["litscout_status"]["status"])
            self.assertEqual("FileNotFoundError", run["litscout_status"]["error_type"])
            self.assertEqual([], run["append_agent_suggestions"])
            self.assertEqual(1, report["summary"]["litscout_failures"])
            self.assertEqual(0, report["summary"]["suggestion_rows_to_append"])

    def test_agent_report_records_litscout_command_failure_without_suggestion(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            failure = subprocess.CalledProcessError(2, ["litscout", "search", "multi"])
            with patch("lab_notebook_agent.agent.subprocess.run", side_effect=failure):
                report = build_agent_report(
                    load_workbook_tables(workbook_path),
                    AgentRunConfig(
                        experiment_ids=("EP-001",),
                        run_litscout=True,
                        artifacts_dir=tmpdir,
                    ),
                )

            run = report["runs"][0]
            self.assertEqual("skipped", run["status"])
            self.assertEqual("litscout_failed", run["skip_reason"])
            self.assertEqual(2, run["litscout_status"]["returncode"])
            self.assertEqual("litscout search multi", run["litscout_status"]["command"])
            self.assertEqual(0, report["summary"]["suggestion_rows_to_append"])

    def test_notebook_context_filters_current_experiment_rows(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tables = load_workbook_tables(save_workbook(Path(tmpdir) / "template.xlsx"))
            matches = notebook_context_matches(
                tables,
                "emulsion polymerization surfactant particle size coagulum",
                "EP-001",
                limit=5,
            )
            self.assertTrue(matches)
            self.assertFalse(
                any(row["sheet"] in {"Experiments", "Daily Log", "Formulations", "Results"} for row in matches),
                matches,
            )

    def test_agent_report_can_disable_notebook_context(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            works_path = write_fake_litscout_export(Path(tmpdir) / "works.json")
            report = build_agent_report(
                load_workbook_tables(workbook_path),
                AgentRunConfig(experiment_ids=("EP-001",), litscout_export=str(works_path), context_limit=0),
            )
            self.assertEqual([], report["runs"][0]["notebook_context_matches"])

    def test_agent_config_sheet_controls_defaults_and_litscout_commands(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            works_path = write_fake_litscout_export(Path(tmpdir) / "works.json")
            tables = load_workbook_tables(workbook_path)
            set_agent_config(tables, "default_context_limit", "0")
            set_agent_config(tables, "default_litscout_sources", "crossref")
            set_agent_config(tables, "default_litscout_depth", "medium")
            set_agent_config(tables, "default_litscout_limit", "2")

            report = build_agent_report(
                tables,
                AgentRunConfig(experiment_ids=("EP-001",), litscout_export=str(works_path)),
            )

            self.assertEqual(0, report["agent_config"]["effective_config"]["context_limit"])
            self.assertEqual(2, report["agent_config"]["effective_config"]["litscout_limit"])
            self.assertEqual("crossref", report["agent_config"]["applied_overrides"]["litscout_sources"])
            self.assertEqual([], report["runs"][0]["notebook_context_matches"])
            command = report["runs"][0]["append_agent_suggestions"][0]["litscout"]["commands"][0]
            self.assertIn("--sources crossref", command)
            self.assertIn("--depth medium --limit 2", command)

    def test_agent_config_sheet_does_not_override_explicit_runtime_values(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            works_path = write_fake_litscout_export(Path(tmpdir) / "works.json")
            tables = load_workbook_tables(workbook_path)
            set_agent_config(tables, "default_context_limit", "0")

            report = build_agent_report(
                tables,
                AgentRunConfig(experiment_ids=("EP-001",), litscout_export=str(works_path), context_limit=2),
            )

            self.assertEqual(2, report["agent_config"]["effective_config"]["context_limit"])
            self.assertNotIn("context_limit", report["agent_config"]["applied_overrides"])
            self.assertTrue(report["runs"][0]["notebook_context_matches"])

    def test_agent_can_require_literature_evidence_before_suggestion(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tables = load_workbook_tables(save_workbook(Path(tmpdir) / "template.xlsx"))

            report = build_agent_report(
                tables,
                AgentRunConfig(experiment_ids=("EP-001",), require_literature_evidence=True),
            )

            run = report["runs"][0]
            self.assertEqual("skipped", run["status"])
            self.assertEqual("literature_evidence_required", run["skip_reason"])
            self.assertEqual(1, report["summary"]["literature_evidence_required"])
            self.assertEqual(0, report["summary"]["suggestion_rows_to_append"])
            self.assertEqual([], run["append_agent_suggestions"])

    def test_agent_config_can_require_literature_evidence(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tables = load_workbook_tables(save_workbook(Path(tmpdir) / "template.xlsx"))
            set_agent_config(tables, "require_literature_evidence", "true")

            report = build_agent_report(tables, AgentRunConfig(experiment_ids=("EP-001",)))

            self.assertTrue(report["agent_config"]["effective_config"]["require_literature_evidence"])
            self.assertTrue(report["agent_config"]["applied_overrides"]["require_literature_evidence"])
            self.assertEqual("literature_evidence_required", report["runs"][0]["skip_reason"])

    def test_required_literature_evidence_allows_litscout_exported_evidence(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            works_path = write_fake_litscout_export(Path(tmpdir) / "works.json")

            report = build_agent_report(
                load_workbook_tables(workbook_path),
                AgentRunConfig(
                    experiment_ids=("EP-001",),
                    litscout_export=str(works_path),
                    require_literature_evidence=True,
                ),
            )

            run = report["runs"][0]
            self.assertEqual("ready", run["status"])
            self.assertEqual(1, report["summary"]["evidence_rows_to_append"])
            self.assertEqual(1, report["summary"]["suggestion_rows_to_append"])
            self.assertEqual(["LIT-EP-001-001"], run["append_agent_suggestions"][0]["linked_evidence_ids"])

    def test_litscout_prediction_report_separates_evidence_inference_and_missing_skills(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            works_path = write_fake_litscout_export(Path(tmpdir) / "works.json")

            report = build_litscout_prediction_report(
                load_workbook_tables(workbook_path),
                AgentRunConfig(experiment_ids=("EP-001",), litscout_export=str(works_path)),
            )

            self.assertEqual("lab-notebook-litscout-prediction.v1", report["schema"])
            prediction = report["predictions"][0]
            self.assertEqual("blocked", prediction["status"])
            self.assertEqual("EP-001-FUP-001", prediction["prediction"]["suggested_experiment_id"])
            self.assertEqual(["LIT-EP-001-001"], prediction["evidence"]["selected_literature_evidence_ids"])
            self.assertTrue(prediction["evidence"]["supporting_findings"])
            self.assertIn("rationale", prediction["inference"])
            self.assertIn("acceptance_criteria", prediction["go_no_go"])
            gap_codes = {gap["code"] for gap in prediction["missing_skill_set"]}
            self.assertIn("safety_review_required", gap_codes)
            self.assertIn("litscout_evidence_unreviewed", gap_codes)
            self.assertIn("formulation_quantities_missing", gap_codes)
            self.assertIn("reagent_properties_missing", gap_codes)
            self.assertIn("result_metrics_missing", gap_codes)
            self.assertEqual(1, report["summary"]["prediction_count"])
            self.assertEqual(1, report["summary"]["prediction_blocker_count"])

    def test_litscout_prediction_report_blocks_missing_or_failed_safety_check(self) -> None:
        base_suggestion = {
            "suggestion_id": "SUG-SAFE-001",
            "confidence": "medium",
            "proposed_change": "Run a controlled follow-up.",
            "expected_effect": "The next result should be attributable.",
            "literature_context": {},
            "proposed_experiment_plan": {
                "suggested_experiment_id": "SAFE-001-FUP-001",
                "process_type": "screening",
                "acceptance_criteria": ["Safety review is complete."],
            },
        }
        for safety_check, expected_code in (
            ("", "safety_check_missing_or_failed"),
            ("Safety review failed: do not run.", "safety_check_missing_or_failed"),
        ):
            run = {
                "experiment_id": "SAFE-001",
                "status": "ready",
                "selected_literature_evidence_ids": ["LIT-SAFE-001"],
                "append_agent_suggestions": [dict(base_suggestion, safety_check=safety_check)],
            }

            prediction = prediction_from_agent_run({"Results": []}, run, {"checks": []})

            self.assertEqual("blocked", prediction["status"])
            gap_codes = {gap["code"] for gap in prediction["missing_skill_set"]}
            self.assertIn(expected_code, gap_codes)

    def test_litscout_prediction_report_allows_approved_safety_check(self) -> None:
        run = {
            "experiment_id": "SAFE-002",
            "status": "ready",
            "selected_literature_evidence_ids": ["LIT-SAFE-002"],
            "append_agent_suggestions": [
                {
                    "suggestion_id": "SUG-SAFE-002",
                    "confidence": "medium",
                    "proposed_change": "Run a controlled follow-up.",
                    "expected_effect": "The next result should be attributable.",
                    "safety_check": "Safety review passed: SDS reviewed and SOP reviewed before execution.",
                    "historical_context": {"prior_experiment_count": 1},
                    "literature_context": {},
                    "proposed_experiment_plan": {
                        "suggested_experiment_id": "SAFE-002-FUP-001",
                        "process_type": "screening",
                        "acceptance_criteria": ["Safety review is complete."],
                    },
                }
            ],
        }

        prediction = prediction_from_agent_run({"Results": []}, run, {"checks": []})

        self.assertEqual("predicted", prediction["status"])
        self.assertEqual([], prediction["missing_skill_set"])

    def test_litscout_prediction_report_blocks_when_grounding_is_required_but_missing(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")

            report = build_litscout_prediction_report(
                load_workbook_tables(workbook_path),
                AgentRunConfig(experiment_ids=("EP-001",), require_literature_evidence=True),
            )

            prediction = report["predictions"][0]
            self.assertEqual("blocked", prediction["status"])
            gap_codes = {gap["code"] for gap in prediction["missing_skill_set"]}
            self.assertIn("suggestion_missing", gap_codes)
            self.assertIn("literature_evidence_missing", gap_codes)
            self.assertEqual(1, report["summary"]["prediction_blocker_count"])

    def test_predict_next_experiment_cli_writes_prediction_report(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            works_path = write_fake_litscout_export(Path(tmpdir) / "works.json")
            output_path = Path(tmpdir) / "prediction.json"

            with patch("builtins.print"):
                exit_code = main(
                    [
                        "predict-next-experiment",
                        "--workbook",
                        str(workbook_path),
                        "--experiment-id",
                        "EP-001",
                        "--litscout-export",
                        str(works_path),
                        "--output",
                        str(output_path),
                    ]
                )

            report = json.loads(output_path.read_text(encoding="utf-8"))
            self.assertEqual(0, exit_code)
            self.assertEqual("lab-notebook-litscout-prediction.v1", report["schema"])
            self.assertEqual("blocked", report["predictions"][0]["status"])

    def test_agent_run_cli_can_require_literature_evidence(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            output = Path(tmpdir) / "agent-report.json"

            with patch("builtins.print"):
                exit_code = main(
                    [
                        "agent-run",
                        "--workbook",
                        str(workbook_path),
                        "--experiment-id",
                        "EP-001",
                        "--require-literature-evidence",
                        "--report-output",
                        str(output),
                    ]
                )

            report = json.loads(output.read_text(encoding="utf-8"))
            self.assertEqual(0, exit_code)
            self.assertEqual("literature_evidence_required", report["runs"][0]["skip_reason"])
            self.assertEqual(1, report["summary"]["literature_evidence_required"])

    def test_agent_run_cli_can_allow_ungrounded_suggestions_over_agent_config(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            workbook = load_workbook(workbook_path)
            config_sheet = workbook["Agent Config"]
            for row in config_sheet.iter_rows(min_row=2):
                if row[0].value == "require_literature_evidence":
                    row[1].value = "true"
                    break
            workbook.save(workbook_path)
            output = Path(tmpdir) / "agent-report.json"

            with patch("builtins.print"):
                exit_code = main(
                    [
                        "agent-run",
                        "--workbook",
                        str(workbook_path),
                        "--experiment-id",
                        "EP-001",
                        "--allow-ungrounded-suggestions",
                        "--report-output",
                        str(output),
                    ]
                )

            report = json.loads(output.read_text(encoding="utf-8"))
            self.assertEqual(0, exit_code)
            self.assertEqual("ready", report["runs"][0]["status"])
            self.assertFalse(report["agent_config"]["effective_config"]["require_literature_evidence"])
            self.assertNotIn("require_literature_evidence", report["agent_config"]["applied_overrides"])
            self.assertEqual(1, report["summary"]["suggestion_rows_to_append"])

    def test_agent_config_confidence_floor_skips_low_confidence_suggestion(self) -> None:
        tables = low_confidence_agent_tables(confidence_floor="medium")

        report = build_agent_report(tables, AgentRunConfig(experiment_ids=("GEN-001",)))

        run = report["runs"][0]
        self.assertEqual("skipped", run["status"])
        self.assertEqual("suggestion_confidence_below_floor", run["skip_reason"])
        self.assertEqual("low", run["suggestion_confidence"])
        self.assertEqual("medium", run["suggestion_confidence_floor"])
        self.assertEqual(0, report["summary"]["suggestion_rows_to_append"])
        self.assertEqual(1, report["summary"]["confidence_below_floor"])
        self.assertEqual([], run["append_agent_suggestions"])
        self.assertEqual("low", run["suppressed_suggestion"]["confidence"])

    def test_agent_config_low_confidence_floor_allows_low_confidence_suggestion(self) -> None:
        tables = low_confidence_agent_tables(confidence_floor="low")

        report = build_agent_report(tables, AgentRunConfig(experiment_ids=("GEN-001",)))

        run = report["runs"][0]
        self.assertEqual("ready", run["status"])
        self.assertEqual("low", run["suggestion_confidence"])
        self.assertEqual(1, report["summary"]["suggestion_rows_to_append"])
        self.assertEqual("low", run["append_agent_suggestions"][0]["confidence"])

    def test_runtime_confidence_floor_overrides_agent_config_floor(self) -> None:
        tables = low_confidence_agent_tables(confidence_floor="medium")

        report = build_agent_report(
            tables,
            AgentRunConfig(
                experiment_ids=("GEN-001",),
                suggestion_confidence_floor="low",
            ),
        )

        run = report["runs"][0]
        self.assertEqual("ready", run["status"])
        self.assertEqual("low", run["suggestion_confidence_floor"])
        self.assertEqual("low", report["agent_config"]["effective_config"]["suggestion_confidence_floor"])
        self.assertNotIn("suggestion_confidence_floor", report["agent_config"]["applied_overrides"])
        self.assertEqual(1, report["summary"]["suggestion_rows_to_append"])

    def test_agent_run_cli_confidence_floor_overrides_agent_config_floor(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = write_tables_to_workbook(
                Path(tmpdir) / "low-confidence.xlsx",
                low_confidence_agent_tables(confidence_floor="medium"),
            )
            output = Path(tmpdir) / "agent-report.json"

            with patch("builtins.print"):
                exit_code = main(
                    [
                        "agent-run",
                        "--workbook",
                        str(workbook_path),
                        "--experiment-id",
                        "GEN-001",
                        "--suggestion-confidence-floor",
                        "low",
                        "--report-output",
                        str(output),
                    ]
                )

            report = json.loads(output.read_text(encoding="utf-8"))
            run = report["runs"][0]
            self.assertEqual(0, exit_code)
            self.assertEqual("ready", run["status"])
            self.assertEqual("low", run["suggestion_confidence_floor"])
            self.assertEqual("low", report["agent_config"]["effective_config"]["suggestion_confidence_floor"])
            self.assertNotIn("suggestion_confidence_floor", report["agent_config"]["applied_overrides"])
            self.assertEqual(1, report["summary"]["suggestion_rows_to_append"])

    def test_daily_review_selects_experiments_by_date_and_log_timestamp(self) -> None:
        tables = {
            "Experiments": [
                {"experiment_id": "EP-001", "date": "2026-06-09", "status": "complete"},
                {"experiment_id": "EP-002", "date": "2026-06-08", "status": "complete"},
                {"experiment_id": "EP-003", "date": "2026-06-08", "status": "running"},
                {"experiment_id": "EP-004", "date": "2026-06-09", "status": "abandoned"},
            ],
            "Daily Log": [
                {"experiment_id": "EP-003", "timestamp": "2026-06-09T15:30:00"},
                {"experiment_id": "EP-002", "timestamp": "2026-06-08T09:00:00"},
                {"experiment_id": "EP-004", "timestamp": "2026-06-09T09:00:00"},
            ],
        }
        self.assertEqual(
            ["EP-001", "EP-003"],
            selected_experiment_ids(tables, (), review_date="2026-06-09"),
        )

    def test_agent_report_records_daily_review_selection(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            works_path = write_fake_litscout_export(Path(tmpdir) / "works.json")
            report = build_agent_report(
                load_workbook_tables(workbook_path),
                AgentRunConfig(review_date="2026-06-09", litscout_export=str(works_path)),
            )
            self.assertEqual("2026-06-09", report["selection"]["review_date"])
            self.assertEqual(["EP-001"], report["selection"]["selected_experiment_ids"])
            self.assertEqual(1, report["summary"]["ready"])

    def test_daily_summary_reports_observations_results_and_material_actions(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            report = build_daily_summary_report(
                load_workbook_tables(workbook_path),
                review_date="2026-06-09",
            )
            self.assertEqual("2026-06-09", report["review_date"])
            self.assertEqual(["EP-001"], report["selection"]["selected_experiment_ids"])
            self.assertEqual(1, report["summary"]["experiment_count"])
            self.assertEqual(1, report["summary"]["observation_count"])
            self.assertEqual(1, report["summary"]["result_count"])
            experiment = report["experiments"][0]
            self.assertIn("coagulum", experiment["issue_tags"])
            self.assertIn("particle_size_high", experiment["result_signals"])
            self.assertEqual(["EP-001"], report["summary"]["experiments_with_result_limits"])
            self.assertEqual(1, report["summary"]["result_limiting_metric_count"])
            self.assertEqual("particle_size", experiment["limiting_metrics"][0]["metric_key"])
            self.assertIn("Outcome limits", experiment["result_analysis_summary"])
            self.assertTrue(
                any("Review result limits" in action for action in experiment["next_actions"]),
                experiment["next_actions"],
            )
            self.assertFalse(experiment["ready_for_quantitative_suggestion"])
            self.assertTrue(experiment["material_recommendations"])

    def test_daily_summary_includes_open_suggestions(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            works_path = write_fake_litscout_export(Path(tmpdir) / "works.json")
            applied_path = Path(tmpdir) / "applied.xlsx"
            run_workbook_agent(
                workbook_path,
                AgentRunConfig(experiment_ids=("EP-001",), litscout_export=str(works_path)),
                apply=True,
                output_workbook=applied_path,
            )
            report = build_daily_summary_report(
                load_workbook_tables(applied_path),
                review_date="2026-06-09",
            )
            self.assertEqual(1, report["summary"]["open_suggestion_count"])
            self.assertEqual("draft", report["experiments"][0]["open_suggestions"][0]["status"])

    def test_daily_summary_guides_completed_planned_suggestion_closure(self) -> None:
        report = build_daily_summary_report(
            {
                "Experiments": [
                    {
                        "experiment_id": "EP-001",
                        "date": "2026-06-09",
                        "process_type": "emulsion polymerization",
                        "status": "complete",
                    },
                    {
                        "experiment_id": "EP-001-FUP-001",
                        "date": "2026-06-10",
                        "process_type": "emulsion polymerization",
                        "status": "complete",
                    },
                ],
                "Formulations": [],
                "Master Reagents": [],
                "Daily Log": [],
                "Results": [],
                "Agent Suggestions": [
                    {
                        "suggestion_id": "SUG-EP-001",
                        "experiment_id": "EP-001",
                        "recommendation_type": "next_experiment",
                        "proposed_experiment_id": "EP-001-FUP-001",
                        "confidence": "medium",
                        "status": "run_planned",
                    }
                ],
            },
            review_date="2026-06-09",
        )
        experiment = report["experiments"][0]
        self.assertEqual("complete", experiment["open_suggestions"][0]["proposed_experiment_status"])
        self.assertTrue(
            any("run_complete" in action and "SUG-EP-001" in action for action in experiment["next_actions"]),
            experiment["next_actions"],
        )

    def test_daily_agent_emits_run_complete_update_for_completed_followup(self) -> None:
        tables = {
            "Experiments": [
                {
                    "experiment_id": "EP-001",
                    "date": "2026-06-09",
                    "process_type": "emulsion polymerization",
                    "status": "complete",
                },
                {
                    "experiment_id": "EP-001-FUP-001",
                    "date": "2026-06-10",
                    "process_type": "emulsion polymerization",
                    "status": "complete",
                },
            ],
            "Formulations": [],
            "Master Reagents": [],
            "Daily Log": [],
            "Results": [],
            "Literature Evidence": [],
            "Agent Suggestions": [
                {
                    "suggestion_id": "SUG-EP-001",
                    "experiment_id": "EP-001",
                    "recommendation_type": "next_experiment",
                    "proposed_experiment_id": "EP-001-FUP-001",
                    "confidence": "medium",
                    "status": "run_planned",
                }
            ],
            "Daily Reviews": [],
            "Process Knowledge": [],
            "Controlled Vocab": [],
            "Agent Config": [],
        }
        snapshot = snapshot_from_tables(
            tables,
            {"Experiments": 101, "Agent Suggestions": 222, "Daily Reviews": 333},
        )
        run = build_snapshot_daily_agent_run(
            snapshot,
            AgentRunConfig(review_date="2026-06-09"),
        )
        self.assertTrue(run["apply_audit"]["valid"], run["apply_audit"])
        self.assertEqual(1, run["summary"]["suggestion_rows_to_update"])
        self.assertEqual(1, run["apply_report"]["summary"]["suggestion_rows_to_update"])
        update = run["update_agent_suggestions"][0]
        self.assertEqual("SUG-EP-001", update["suggestion_id"])
        self.assertEqual("run_complete", update["value"])
        self.assertEqual(1, run["apply_audit"]["summary"]["suggestion_rows_to_update"])
        request = run["batch_update_requests"][-1]
        self.assertEqual(222, request["updateCells"]["start"]["sheetId"])
        agent_suggestion_headers = list(next(spec for spec in SHEETS if spec.name == "Agent Suggestions").headers)
        self.assertEqual(agent_suggestion_headers.index("status"), request["updateCells"]["start"]["columnIndex"])
        self.assertEqual(
            "run_complete",
            request["updateCells"]["rows"][0]["values"][0]["userEnteredValue"]["stringValue"],
        )

    def test_daily_log_results_normalizes_structured_observations(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            report = build_daily_log_results_report(
                load_workbook_tables(workbook_path),
                experiment_ids=("EP-001",),
            )
            self.assertEqual("lab-notebook-agent-daily-log-results.v1", report["schema"])
            self.assertEqual(1, report["summary"]["daily_log_rows_considered"])
            self.assertEqual(2, report["summary"]["result_rows_to_append"])
            self.assertEqual(1, report["summary"]["measurements_skipped"])
            rows = report["runs"][0]["append_results"]
            measurement_types = {row["measurement_type"] for row in rows}
            self.assertEqual({"temperature", "agitation speed"}, measurement_types)
            self.assertEqual("DLS particle size", report["runs"][0]["skipped_measurements"][0]["measurement_type"])

    def test_daily_log_results_extracts_measurements_from_observation_text(self) -> None:
        report = build_daily_log_results_report(
            {
                "Daily Log": [
                    {
                        "experiment_id": "EP-TEXT",
                        "timestamp": "2026-06-09T16:00:00",
                        "process_stage": "test",
                        "particle_size_nm": "510",
                        "observation": (
                            "DLS 510 nm after workup; conversion 76%; pH 5.8; "
                            "solids 39%; viscosity 120 cP; coagulum mass 0.4 g."
                        ),
                    }
                ],
                "Results": [],
            },
            experiment_ids=("EP-TEXT",),
        )
        rows = report["runs"][0]["append_results"]
        by_type = {row["measurement_type"]: row for row in rows}
        self.assertEqual(6, report["summary"]["result_rows_to_append"])
        self.assertEqual("510", by_type["DLS particle size"]["value"])
        self.assertEqual("Daily Log structured field", by_type["DLS particle size"]["method"])
        self.assertEqual("76", by_type["conversion"]["value"])
        self.assertEqual("5.8", by_type["pH"]["value"])
        self.assertEqual("39", by_type["solids percent"]["value"])
        self.assertEqual("120", by_type["viscosity"]["value"])
        self.assertEqual("0.4", by_type["coagulum mass"]["value"])
        self.assertTrue(all("observation text" in row["method"] for name, row in by_type.items() if name != "DLS particle size"))

    def test_daily_log_results_extracts_polymer_specific_outcomes_from_text(self) -> None:
        report = build_daily_log_results_report(
            {
                "Daily Log": [
                    {
                        "experiment_id": "EP-POLY",
                        "timestamp": "2026-06-09T18:30:00",
                        "process_stage": "post-test",
                        "observation": (
                            "Residual monomer 1.8%; PDI 0.12; Tg -18 C; "
                            "thermal hold for 45 min."
                        ),
                    }
                ],
                "Results": [],
            },
            experiment_ids=("EP-POLY",),
        )

        rows = report["runs"][0]["append_results"]
        by_type = {row["measurement_type"]: row for row in rows}
        self.assertEqual(4, report["summary"]["result_rows_to_append"])
        self.assertEqual("1.8", by_type["residual monomer"]["value"])
        self.assertEqual("%", by_type["residual monomer"]["units"])
        self.assertEqual("0.12", by_type["polydispersity index"]["value"])
        self.assertEqual("-18", by_type["Tg"]["value"])
        self.assertEqual("C", by_type["Tg"]["units"])
        self.assertEqual("45", by_type["hold time"]["value"])
        self.assertEqual("min", by_type["hold time"]["units"])

    def test_daily_log_results_normalizes_structured_polymer_outcomes(self) -> None:
        report = build_daily_log_results_report(
            {
                "Daily Log": [
                    {
                        "experiment_id": "EP-STRUCT",
                        "timestamp": "2026-06-09T19:00:00",
                        "process_stage": "test",
                        "residual_monomer_percent": "1.8",
                        "polydispersity_index": "0.12",
                        "Tg_C": "-18",
                        "hold_time_min": "45",
                    }
                ],
                "Results": [],
            },
            experiment_ids=("EP-STRUCT",),
        )

        rows = report["runs"][0]["append_results"]
        by_type = {row["measurement_type"]: row for row in rows}
        self.assertEqual(4, report["summary"]["result_rows_to_append"])
        self.assertEqual("Daily Log structured field", by_type["residual monomer"]["method"])
        self.assertEqual("1.8", by_type["residual monomer"]["value"])
        self.assertEqual("0.12", by_type["polydispersity index"]["value"])
        self.assertEqual("-18", by_type["Tg"]["value"])
        self.assertEqual("45", by_type["hold time"]["value"])

    def test_daily_log_results_apply_writes_results_rows_to_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            output_path = Path(tmpdir) / "normalized.xlsx"
            report = build_daily_log_results_report(load_workbook_tables(workbook_path), experiment_ids=("EP-001",))
            apply_daily_log_results_report_to_workbook(workbook_path, report, output_path)
            workbook = load_workbook(output_path)
            worksheet = workbook["Results"]
            self.assertEqual("temperature", worksheet["C3"].value)
            self.assertEqual("70", worksheet["E3"].value)
            self.assertEqual("agitation speed", worksheet["C4"].value)

    def test_daily_log_results_snapshot_emits_google_batch_requests(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            snapshot = snapshot_from_tables(
                load_workbook_tables(workbook_path),
                {"Results": 102},
            )
            report = build_daily_log_results_report(
                snapshot_to_tables(snapshot),
                experiment_ids=("EP-001",),
            )
            audit = audit_report_against_snapshot(report, snapshot)
            self.assertTrue(audit["valid"], audit["errors"])
            self.assertEqual(2, audit["summary"]["result_rows_to_append"])
            requests = batch_update_requests_from_report(report, sheet_ids_from_snapshot(snapshot))
            self.assertEqual(1, len(requests))
            self.assertEqual(102, requests[0]["appendCells"]["sheetId"])

    def test_daily_agent_run_combines_summary_and_agent_report(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            works_path = write_fake_litscout_export(Path(tmpdir) / "works.json")
            run = build_daily_agent_run(
                load_workbook_tables(workbook_path),
                AgentRunConfig(review_date="2026-06-09", litscout_export=str(works_path)),
            )
            self.assertEqual("lab-notebook-agent-daily-run.v1", run["schema"])
            self.assertEqual("2026-06-09", run["review_date"])
            self.assertEqual(["EP-001"], run["selection"]["selected_experiment_ids"])
            self.assertEqual(1, run["daily_summary"]["summary"]["experiment_count"])
            self.assertEqual(1, run["agent_report"]["summary"]["suggestion_rows_to_append"])
            self.assertEqual(1, run["summary"]["evidence_rows_to_append"])
            self.assertEqual(2, run["summary"]["normalized_result_rows_to_append"])
            self.assertEqual("lab-notebook-agent-daily-log-results.v1", run["daily_log_results_report"]["schema"])
            self.assertEqual(1, run["summary"]["experiment_review_count"])
            self.assertEqual(2, run["summary"]["preflight_fail_count"])
            self.assertEqual(1, run["summary"]["result_limiting_metric_count"])
            self.assertEqual(["EP-001"], run["summary"]["experiments_with_result_limits"])
            review = run["experiment_reviews"][0]
            self.assertEqual("EP-001", review["experiment_id"])
            self.assertEqual("lab-notebook-agent-experiment-preflight.v1", review["preflight"]["schema"])
            self.assertEqual("lab-notebook-agent-process-material-search.v1", review["material_search"]["schema"])
            roles = {role["role_group"]: role for role in review["material_search"]["roles"]}
            self.assertEqual("M-SKA", roles["monomer"]["candidate_reagents"][0]["reagent_id"])

    def test_daily_agent_projects_normalized_daily_log_results_into_suggestions(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            tables = load_workbook_tables(workbook_path)
            complete_template_materials(tables)
            tables["Results"] = []

            run = build_daily_agent_run(
                tables,
                AgentRunConfig(review_date="2026-06-09"),
            )

            self.assertEqual(3, run["summary"]["normalized_result_rows_to_append"])
            suggestion = run["agent_report"]["runs"][0]["append_agent_suggestions"][0]
            self.assertIn("particle_size_high", suggestion["result_analysis"]["signals"])
            self.assertTrue(suggestion["proposed_experiment_plan"]["result_support"]["limiting_metrics"])
            checks = {row["name"]: row for row in run["experiment_reviews"][0]["preflight"]["checks"]}
            self.assertEqual("pass", checks["results_measurements"]["status"])

    def test_daily_agent_projects_normalized_formulations_into_suggestions(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            tables = load_workbook_tables(workbook_path)
            for row in tables["Master Reagents"]:
                if row["reagent_id"] == "M-SKA":
                    row["molecular_weight_g_mol"] = "100"
                    row["density_g_mL"] = "1.25"
                    row["purity_fraction"] = "0.8"
            for row in tables["Formulations"]:
                if row["reagent_id"] == "M-SKA":
                    row["mass_g"] = "10"

            run = build_daily_agent_run(
                tables,
                AgentRunConfig(review_date="2026-06-09"),
            )

            self.assertEqual(2, run["summary"]["formulation_cells_to_update"])
            updates = {
                update["field"]: update["value"]
                for report_run in run["formulation_normalization_report"]["runs"]
                if report_run["reagent_id"] == "M-SKA"
                for update in report_run["update_formulations"]
            }
            self.assertEqual({"volume_mL": "8", "moles_mmol": "80"}, updates)
            suggestion = run["agent_report"]["runs"][0]["append_agent_suggestions"][0]
            formulation = next(
                row
                for row in suggestion["proposed_experiment_plan"]["sheet_rows"]["formulations"]
                if row["reagent_id"] == "M-SKA"
            )
            self.assertEqual("8", formulation["volume_mL"])
            self.assertEqual("80", formulation["moles_mmol"])
            self.assertTrue(
                any(
                    update["field"] == "summary" and "2 normalized Formulations pending" in update["value"]
                    for update in run["update_experiments"]
                )
            )

    def test_daily_review_row_includes_result_limit_action(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            works_path = write_fake_litscout_export(Path(tmpdir) / "works.json")
            run = build_daily_agent_run(
                load_workbook_tables(workbook_path),
                AgentRunConfig(review_date="2026-06-09", litscout_export=str(works_path)),
            )
            review_row = daily_review_row_from_run(run)
            actions = json.loads(review_row["next_actions_json"])
            self.assertTrue(
                any(action.startswith("Review result limits for EP-001") for action in actions),
                actions,
            )

    def test_daily_review_row_counts_formulation_updates_as_apply_work(self) -> None:
        review_row = daily_review_row_from_run(
            {
                "review_date": "2026-06-09",
                "selection": {"selected_experiment_ids": ["EP-001"]},
                "summary": {
                    "experiment_count": 1,
                    "observation_count": 0,
                    "result_count": 0,
                    "normalized_result_rows_to_append": 0,
                    "formulation_cells_to_update": 2,
                    "evidence_rows_to_append": 0,
                    "suggestion_rows_to_append": 0,
                    "suggestion_rows_to_update": 0,
                    "preflight_fail_count": 0,
                    "preflight_warn_count": 0,
                },
                "daily_summary": {"experiments": []},
                "experiment_reviews": [],
            }
        )

        actions = json.loads(review_row["next_actions_json"])
        self.assertEqual("ready_to_apply", review_row["status"])
        self.assertIn("2 normalized Formulations cells", review_row["summary"])
        self.assertTrue(
            any("normalized Formulations quantities" in action for action in actions),
            actions,
        )

    def test_snapshot_daily_agent_run_includes_audit_and_batch_requests(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            works_path = write_fake_litscout_export(Path(tmpdir) / "works.json")
            snapshot = snapshot_from_tables(
                load_workbook_tables(workbook_path),
                {"Experiments": 101, "Results": 102, "Literature Evidence": 111, "Agent Suggestions": 222, "Daily Reviews": 333},
            )
            run = build_snapshot_daily_agent_run(
                snapshot,
                AgentRunConfig(review_date="2026-06-09", litscout_export=str(works_path)),
            )
            self.assertTrue(run["snapshot_audit"]["valid"], run["snapshot_audit"])
            self.assertTrue(run["apply_audit"]["valid"], run["apply_audit"])
            self.assertEqual(1, len(run["experiment_reviews"]))
            self.assertEqual(1, run["apply_report"]["summary"]["daily_review_rows_to_append"])
            self.assertEqual(1, run["apply_audit"]["summary"]["daily_review_rows_to_append"])
            self.assertEqual(4, run["summary"]["experiment_cells_to_update"])
            self.assertEqual(4, run["apply_audit"]["summary"]["experiment_cells_to_update"])
            self.assertEqual(8, run["summary"]["apply_request_count"])
            self.assertEqual(102, run["batch_update_requests"][0]["appendCells"]["sheetId"])
            self.assertEqual(111, run["batch_update_requests"][1]["appendCells"]["sheetId"])
            self.assertEqual(222, run["batch_update_requests"][2]["appendCells"]["sheetId"])
            self.assertEqual(333, run["batch_update_requests"][3]["appendCells"]["sheetId"])
            self.assertEqual(101, run["batch_update_requests"][4]["updateCells"]["start"]["sheetId"])

    def test_snapshot_daily_agent_run_batches_formulation_normalization(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            tables = load_workbook_tables(workbook_path)
            for row in tables["Master Reagents"]:
                if row["reagent_id"] == "M-SKA":
                    row["molecular_weight_g_mol"] = "100"
                    row["density_g_mL"] = "1.25"
                    row["purity_fraction"] = "0.8"
            for row in tables["Formulations"]:
                if row["reagent_id"] == "M-SKA":
                    row["mass_g"] = "10"
            snapshot = snapshot_from_tables(
                tables,
                {
                    "Experiments": 101,
                    "Formulations": 103,
                    "Results": 102,
                    "Agent Suggestions": 222,
                    "Daily Reviews": 333,
                },
            )

            run = build_snapshot_daily_agent_run(
                snapshot,
                AgentRunConfig(review_date="2026-06-09"),
            )

            self.assertTrue(run["apply_audit"]["valid"], run["apply_audit"])
            self.assertEqual(2, run["apply_report"]["summary"]["formulation_cells_to_update"])
            self.assertEqual(2, run["apply_audit"]["summary"]["formulation_cells_to_update"])
            formulation_requests = [
                request
                for request in run["batch_update_requests"]
                if request.get("updateCells", {}).get("start", {}).get("sheetId") == 103
            ]
            self.assertEqual([5, 6], [request["updateCells"]["start"]["columnIndex"] for request in formulation_requests])
            self.assertEqual(
                ["8", "80"],
                [
                    request["updateCells"]["rows"][0]["values"][0]["userEnteredValue"]["stringValue"]
                    for request in formulation_requests
                ],
            )

    def test_daily_agent_apply_writes_daily_review_row_to_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            works_path = write_fake_litscout_export(Path(tmpdir) / "works.json")
            output_path = Path(tmpdir) / "daily-applied.xlsx"
            run_workbook_daily_agent(
                workbook_path,
                AgentRunConfig(review_date="2026-06-09", litscout_export=str(works_path)),
                apply=True,
                output_workbook=output_path,
            )
            workbook = load_workbook(output_path)
            self.assertEqual("DRV-20260609-EP001", workbook["Daily Reviews"]["A2"].value)
            self.assertEqual("2026-06-09", workbook["Daily Reviews"]["C2"].value)
            self.assertEqual("needs_attention", workbook["Daily Reviews"]["N2"].value)
            self.assertEqual("LIT-EP-001-001", workbook["Experiments"]["G2"].value)
            self.assertEqual("needs_review", workbook["Experiments"]["I2"].value)
            self.assertIn("Daily review 2026-06-09", workbook["Experiments"]["K2"].value)

    def test_daily_agent_apply_writes_formulation_normalization_to_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            workbook = load_workbook(workbook_path)
            workbook["Master Reagents"]["F2"] = "100"
            workbook["Master Reagents"]["G2"] = "1.25"
            workbook["Master Reagents"]["H2"] = "0.8"
            workbook["Formulations"]["E2"] = "10"
            workbook.save(workbook_path)
            output_path = Path(tmpdir) / "daily-applied.xlsx"

            run_workbook_daily_agent(
                workbook_path,
                AgentRunConfig(review_date="2026-06-09"),
                apply=True,
                output_workbook=output_path,
            )

            workbook = load_workbook(output_path)
            self.assertEqual("8", workbook["Formulations"]["F2"].value)
            self.assertEqual("80", workbook["Formulations"]["G2"].value)

    def test_daily_agent_apply_marks_completed_followup_suggestion_run_complete(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            workbook = load_workbook(workbook_path)
            experiments = workbook["Experiments"]
            experiment_headers = [cell.value for cell in experiments[1]]
            followup = {
                "experiment_id": "EP-001-FUP-001",
                "date": "2026-06-10",
                "project": "latex",
                "process_type": "emulsion polymerization",
                "objective": "Follow-up run",
                "status": "complete",
            }
            experiments.append([followup.get(header, "") for header in experiment_headers])
            suggestions = workbook["Agent Suggestions"]
            suggestion_headers = [cell.value for cell in suggestions[1]]
            suggestion = {
                "suggestion_id": "SUG-EP-001",
                "experiment_id": "EP-001",
                "recommendation_type": "next_experiment",
                "proposed_experiment_id": "EP-001-FUP-001",
                "confidence": "medium",
                "status": "run_planned",
            }
            suggestions.append([suggestion.get(header, "") for header in suggestion_headers])
            workbook.save(workbook_path)

            output_path = Path(tmpdir) / "daily-applied.xlsx"
            run_workbook_daily_agent(
                workbook_path,
                AgentRunConfig(review_date="2026-06-09"),
                apply=True,
                output_workbook=output_path,
            )
            workbook = load_workbook(output_path)
            suggestions = workbook["Agent Suggestions"]
            headers = [cell.value for cell in suggestions[1]]
            self.assertEqual("run_complete", suggestions.cell(row=2, column=headers.index("status") + 1).value)

    def test_agent_run_apply_writes_evidence_and_suggestion_to_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            works_path = write_fake_litscout_export(Path(tmpdir) / "works.json")
            output_path = Path(tmpdir) / "applied.xlsx"
            run_workbook_agent(
                workbook_path,
                AgentRunConfig(experiment_ids=("EP-001",), litscout_export=str(works_path)),
                apply=True,
                output_workbook=output_path,
            )
            workbook = load_workbook(output_path)
            self.assertEqual("LIT-EP-001-001", workbook["Literature Evidence"]["A2"].value)
            self.assertEqual("LIT-EP-001-001", workbook["Experiments"]["G2"].value)
            self.assertEqual("EP-001", workbook["Agent Suggestions"]["C2"].value)
            self.assertEqual("LIT-EP-001-001", workbook["Agent Suggestions"]["H2"].value)
            headers = [cell.value for cell in workbook["Agent Suggestions"][1]]
            self.assertEqual(
                "EP-001-FUP-001",
                workbook["Agent Suggestions"].cell(row=2, column=headers.index("proposed_experiment_id") + 1).value,
            )

    def test_agent_report_skips_existing_suggestion_without_force(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            works_path = write_fake_litscout_export(Path(tmpdir) / "works.json")
            output_path = Path(tmpdir) / "applied.xlsx"
            run_workbook_agent(
                workbook_path,
                AgentRunConfig(experiment_ids=("EP-001",), litscout_export=str(works_path)),
                apply=True,
                output_workbook=output_path,
            )
            report = run_workbook_agent(
                output_path,
                AgentRunConfig(experiment_ids=("EP-001",), litscout_export=str(works_path)),
            )
            self.assertEqual("skipped", report["runs"][0]["status"])
            self.assertEqual("existing_suggestion", report["runs"][0]["skip_reason"])

    def test_agent_report_allows_new_suggestion_after_run_complete(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            works_path = write_fake_litscout_export(Path(tmpdir) / "works.json")
            output_path = Path(tmpdir) / "applied.xlsx"
            run_workbook_agent(
                workbook_path,
                AgentRunConfig(experiment_ids=("EP-001",), litscout_export=str(works_path)),
                apply=True,
                output_workbook=output_path,
            )
            workbook = load_workbook(output_path)
            suggestions = workbook["Agent Suggestions"]
            headers = [cell.value for cell in suggestions[1]]
            suggestions.cell(row=2, column=headers.index("status") + 1, value="run_complete")
            workbook.save(output_path)

            report = run_workbook_agent(
                output_path,
                AgentRunConfig(experiment_ids=("EP-001",), litscout_export=str(works_path)),
            )
            self.assertEqual("ready", report["runs"][0]["status"])
            self.assertEqual(1, report["summary"]["suggestion_rows_to_append"])
            self.assertEqual("existing_evidence", report["runs"][0]["litscout_status"]["status"])
            suggestion = report["runs"][0]["append_agent_suggestions"][0]
            self.assertEqual("EP-001-FUP-002", suggestion["proposed_experiment_plan"]["suggested_experiment_id"])

    def test_next_followup_experiment_id_uses_existing_experiments_and_suggestions(self) -> None:
        tables = {
            "Experiments": [
                {"experiment_id": "EP-001"},
                {"experiment_id": "EP-001-FUP-001"},
                {"experiment_id": "EP-001-FUP-003"},
                {"experiment_id": "EP-002-FUP-009"},
            ],
            "Agent Suggestions": [
                {
                    "suggestion_id": "SUG-001",
                    "experiment_id": "EP-001",
                    "proposed_experiment_id": "EP-001-FUP-002",
                    "status": "run_complete",
                },
                {
                    "suggestion_id": "SUG-002",
                    "experiment_id": "EP-001",
                    "proposed_plan_json": json.dumps({"suggested_experiment_id": "EP-001-FUP-004"}),
                    "status": "rejected",
                },
            ],
        }
        self.assertEqual("EP-001-FUP-005", next_followup_experiment_id(tables, "EP-001"))

    def test_google_batch_requests_from_agent_report(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            works_path = write_fake_litscout_export(Path(tmpdir) / "works.json")
            report = run_workbook_agent(
                workbook_path,
                AgentRunConfig(experiment_ids=("EP-001",), litscout_export=str(works_path)),
            )
            requests = batch_update_requests_from_report(
                report,
                {"Experiments": 101, "Literature Evidence": 111, "Agent Suggestions": 222},
            )
            self.assertEqual(3, len(requests))
            self.assertEqual(111, requests[0]["appendCells"]["sheetId"])
            self.assertEqual(222, requests[1]["appendCells"]["sheetId"])
            self.assertEqual(101, requests[2]["updateCells"]["start"]["sheetId"])
            self.assertEqual(6, requests[2]["updateCells"]["start"]["columnIndex"])
            self.assertEqual("EP-001", requests[1]["appendCells"]["rows"][0]["values"][2]["userEnteredValue"]["stringValue"])

    def test_snapshot_round_trip_drives_agent_report_and_batch(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            works_path = write_fake_litscout_export(Path(tmpdir) / "works.json")
            tables = load_workbook_tables(workbook_path)
            snapshot = snapshot_from_tables(
                tables,
                {"Experiments": 101, "Literature Evidence": 111, "Agent Suggestions": 222},
            )
            report = build_agent_report(
                snapshot_to_tables(snapshot),
                AgentRunConfig(experiment_ids=("EP-001",), litscout_export=str(works_path)),
            )
            requests = batch_update_requests_from_report(report, sheet_ids_from_snapshot(snapshot))
            self.assertEqual(1, report["summary"]["suggestion_rows_to_append"])
            self.assertEqual(3, len(requests))
            self.assertEqual("LIT-EP-001-001", requests[0]["appendCells"]["rows"][0]["values"][0]["userEnteredValue"]["stringValue"])

    def test_parse_sheet_id_args(self) -> None:
        self.assertEqual(
            {"Literature Evidence": 111, "Agent Suggestions": 222},
            parse_sheet_id_args(["Literature Evidence=111", "Agent Suggestions=222"]),
        )

    def test_google_capture_plan_lists_all_contract_sheets(self) -> None:
        plan = snapshot_capture_plan("spreadsheet-123")
        self.assertEqual("spreadsheet-123", plan["spreadsheet_id"])
        self.assertEqual([sheet.name for sheet in SHEETS], [row["sheet_name"] for row in plan["sheets"]])
        apply_sheets = [row["sheet_name"] for row in plan["sheets"] if row["used_for_apply"]]
        self.assertEqual(
            [
                "Master Reagents",
                "Experiments",
                "Daily Log",
                "Formulations",
                "Results",
                "Literature Evidence",
                "Agent Suggestions",
                "Daily Reviews",
            ],
            apply_sheets,
        )

    def test_google_setup_requests_create_tabs_headers_and_validations(self) -> None:
        metadata = {
            "spreadsheetId": "spreadsheet-123",
            "sheets": [
                {
                    "properties": {
                        "title": "Experiments",
                        "sheetId": 101,
                        "gridProperties": {"rowCount": 100, "columnCount": 3},
                    }
                },
                {"properties": {"title": "Notes", "sheetId": 999}},
            ],
        }
        requests = google_setup_requests_from_metadata(metadata)
        added_titles = [
            request["addSheet"]["properties"]["title"]
            for request in requests
            if "addSheet" in request
        ]
        self.assertIn("Master Reagents", added_titles)
        self.assertNotIn("Experiments", added_titles)

        experiment_header_request = next(
            request["updateCells"]
            for request in requests
            if request.get("updateCells", {}).get("start", {}).get("sheetId") == 101
            and request["updateCells"]["start"].get("rowIndex") == 0
        )
        self.assertEqual(
            "experiment_id",
            experiment_header_request["rows"][0]["values"][0]["userEnteredValue"]["stringValue"],
        )
        experiment_grid_request = next(
            request["updateSheetProperties"]
            for request in requests
            if request.get("updateSheetProperties", {}).get("properties", {}).get("sheetId") == 101
        )
        self.assertEqual(
            1000,
            experiment_grid_request["properties"]["gridProperties"]["rowCount"],
        )
        experiment_validations = [
            request["setDataValidation"]
            for request in requests
            if request.get("setDataValidation", {}).get("range", {}).get("sheetId") == 101
        ]
        self.assertTrue(
            any(
                validation["range"]["startColumnIndex"] == 3
                and any(
                    value["userEnteredValue"] == "emulsion polymerization"
                    for value in validation["rule"]["condition"]["values"]
                )
                for validation in experiment_validations
            ),
            experiment_validations,
        )

        audit = google_setup_audit_from_metadata(metadata)
        self.assertEqual(["Notes"], audit["unknown_sheets"])
        self.assertEqual(1, audit["summary"]["existing_contract_sheet_count"])
        self.assertEqual(len(requests), audit["summary"]["request_count"])
        self.assertEqual(900000000, audit["generated_sheet_ids"]["Master Reagents"])

    def test_validate_snapshot_detects_header_mismatch(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tables = load_workbook_tables(save_workbook(Path(tmpdir) / "template.xlsx"))
            snapshot = snapshot_from_tables(tables)
            snapshot["sheets"]["Experiments"]["values"][0][0] = "bad_header"
            audit = validate_snapshot(snapshot)
            self.assertFalse(audit["valid"])
            self.assertEqual("header_mismatch", audit["errors"][0]["code"])

    def test_audit_report_against_snapshot_detects_duplicate_append(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            works_path = write_fake_litscout_export(Path(tmpdir) / "works.json")
            applied_path = Path(tmpdir) / "applied.xlsx"
            report = run_workbook_agent(
                workbook_path,
                AgentRunConfig(experiment_ids=("EP-001",), litscout_export=str(works_path)),
                apply=True,
                output_workbook=applied_path,
            )
            snapshot = snapshot_from_tables(
                load_workbook_tables(applied_path),
                {"Literature Evidence": 111, "Agent Suggestions": 222},
            )
            audit = audit_report_against_snapshot(report, snapshot)
            self.assertFalse(audit["valid"])
            self.assertTrue(any(error["code"] == "duplicate_append" for error in audit["errors"]))

    def test_audit_report_accepts_target_sheet_ids_only(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            works_path = write_fake_litscout_export(Path(tmpdir) / "works.json")
            report = run_workbook_agent(
                workbook_path,
                AgentRunConfig(experiment_ids=("EP-001",), litscout_export=str(works_path)),
            )
            snapshot = snapshot_from_tables(
                load_workbook_tables(workbook_path),
                {"Experiments": 101, "Literature Evidence": 111, "Agent Suggestions": 222},
            )
            audit = audit_report_against_snapshot(report, snapshot)
            self.assertTrue(audit["valid"], audit["errors"])
            self.assertEqual(3, audit["summary"]["request_count"])

    def test_google_api_capture_snapshot_includes_live_sheet_ids(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            tables = load_workbook_tables(save_workbook(Path(tmpdir) / "template.xlsx"))
            client = FakeSheetsApiClient(
                snapshot_from_tables(
                    tables,
                    {
                        "Experiments": 101,
                        "Formulations": 103,
                        "Results": 102,
                        "Literature Evidence": 111,
                        "Agent Suggestions": 222,
                        "Daily Reviews": 333,
                    },
                )
            )
            snapshot = capture_snapshot_from_google_sheets("spreadsheet-123", client)
            self.assertEqual(101, snapshot["sheets"]["Experiments"]["sheet_id"])
            self.assertEqual("experiment_id", snapshot["sheets"]["Experiments"]["values"][0][0])
            self.assertTrue(validate_snapshot(snapshot)["valid"])

    def test_google_api_doctor_reports_contract_ready_with_fake_client(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            client = FakeSheetsApiClient(
                snapshot_from_tables(
                    load_workbook_tables(save_workbook(Path(tmpdir) / "template.xlsx")),
                    {sheet.name: index for index, sheet in enumerate(SHEETS, start=100)},
                )
            )
            with patch("lab_notebook_agent.google_api.GoogleSheetsApiClient.from_credentials", return_value=client):
                result = google_api_doctor("spreadsheet-123")
            self.assertTrue(result["ready"])
            self.assertEqual("passed", result["checks"][-1]["status"])

    def test_google_api_doctor_reports_credential_failure(self) -> None:
        with patch(
            "lab_notebook_agent.google_api.GoogleSheetsApiClient.from_credentials",
            side_effect=RuntimeError("credentials missing"),
        ):
            result = google_api_doctor("spreadsheet-123")
        self.assertFalse(result["ready"])
        self.assertEqual("failed", result["checks"][0]["status"])
        self.assertIn("credentials missing", result["checks"][0]["message"])

    def test_live_google_setup_applies_setup_batch_with_fake_client(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            client = FakeSheetsApiClient(
                snapshot_from_tables(
                    load_workbook_tables(save_workbook(Path(tmpdir) / "template.xlsx")),
                    {sheet.name: index for index, sheet in enumerate(SHEETS, start=100)},
                )
            )
            run = run_live_google_setup("spreadsheet-123", client, apply=True)
            self.assertTrue(run["applied"])
            self.assertEqual("lab-notebook-agent-live-google-setup.v1", run["schema"])
            self.assertEqual(len(run["batch_update_requests"]), run["setup_audit"]["summary"]["request_count"])
            self.assertEqual(1, len(client.batch_updates))
            self.assertFalse(any("addSheet" in request for request in run["batch_update_requests"]))
            self.assertTrue(any("setDataValidation" in request for request in run["batch_update_requests"]))

    def test_live_google_experiment_record_applies_valid_batch_with_fake_client(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            client = FakeSheetsApiClient(
                snapshot_from_tables(
                    load_workbook_tables(save_workbook(Path(tmpdir) / "template.xlsx")),
                    {
                        "Experiments": 101,
                        "Formulations": 103,
                        "Daily Log": 104,
                        "Results": 102,
                    },
                )
            )
            run = run_live_google_experiment_record(
                "spreadsheet-123",
                client,
                sample_experiment_record(),
                apply=True,
            )
            self.assertTrue(run["applied"])
            self.assertEqual("lab-notebook-agent-live-google-experiment-record.v1", run["schema"])
            self.assertEqual("lab-notebook-agent-experiment-record.v1", run["record_report"]["schema"])
            self.assertTrue(run["apply_audit"]["valid"], run["apply_audit"])
            self.assertEqual(
                [101, 103, 104, 102],
                [request["appendCells"]["sheetId"] for request in run["batch_update_requests"]],
            )
            self.assertEqual(1, len(client.batch_updates))

    def test_live_google_recorded_daily_agent_applies_combined_batch_with_fake_client(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            works_path = write_fake_litscout_export(Path(tmpdir) / "works.json")
            client = FakeSheetsApiClient(
                snapshot_from_tables(
                    load_workbook_tables(save_workbook(Path(tmpdir) / "template.xlsx")),
                    {
                        "Experiments": 101,
                        "Formulations": 103,
                        "Daily Log": 104,
                        "Results": 102,
                        "Literature Evidence": 111,
                        "Agent Suggestions": 222,
                        "Daily Reviews": 333,
                    },
                )
            )
            run = run_live_google_recorded_daily_agent(
                "spreadsheet-123",
                client,
                sample_experiment_record(),
                config=AgentRunConfig(litscout_export=str(works_path)),
                apply=True,
            )
            self.assertTrue(run["applied"])
            self.assertEqual("lab-notebook-agent-live-google-recorded-daily-run.v1", run["schema"])
            self.assertEqual("lab-notebook-agent-recorded-daily-run.v1", run["recorded_daily_run"]["schema"])
            self.assertTrue(run["apply_audit"]["valid"], run["apply_audit"])
            self.assertEqual(
                [101, 103, 104, 102, 111, 222, 333],
                [request["appendCells"]["sheetId"] for request in run["batch_update_requests"]],
            )
            self.assertEqual(
                "LIT-EP-010-001",
                run["apply_report"]["append_experiments"][0]["linked_literature_ids"],
            )
            self.assertEqual(1, len(client.batch_updates))

    def test_live_google_agent_run_applies_valid_batch_with_fake_client(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            works_path = write_fake_litscout_export(Path(tmpdir) / "works.json")
            client = FakeSheetsApiClient(
                snapshot_from_tables(
                    load_workbook_tables(workbook_path),
                    {
                        "Experiments": 101,
                        "Results": 102,
                        "Literature Evidence": 111,
                        "Agent Suggestions": 222,
                        "Daily Reviews": 333,
                    },
                )
            )
            run = run_live_google_agent(
                "spreadsheet-123",
                client,
                config=AgentRunConfig(experiment_ids=("EP-001",), litscout_export=str(works_path)),
                apply=True,
            )
            self.assertTrue(run["applied"])
            self.assertTrue(run["apply_audit"]["valid"], run["apply_audit"])
            self.assertEqual(3, len(run["batch_update_requests"]))
            self.assertEqual(1, len(client.batch_updates))
            self.assertEqual(222, run["batch_update_requests"][1]["appendCells"]["sheetId"])
            self.assertEqual(101, run["batch_update_requests"][2]["updateCells"]["start"]["sheetId"])

    def test_live_google_daily_agent_run_applies_valid_batch_with_fake_client(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            works_path = write_fake_litscout_export(Path(tmpdir) / "works.json")
            client = FakeSheetsApiClient(
                snapshot_from_tables(
                    load_workbook_tables(workbook_path),
                    {
                        "Experiments": 101,
                        "Results": 102,
                        "Literature Evidence": 111,
                        "Agent Suggestions": 222,
                        "Daily Reviews": 333,
                    },
                )
            )
            run = run_live_google_daily_agent(
                "spreadsheet-123",
                client,
                config=AgentRunConfig(review_date="2026-06-09", litscout_export=str(works_path)),
                apply=True,
            )
            self.assertTrue(run["applied"])
            self.assertEqual("lab-notebook-agent-daily-run.v1", run["daily_agent_run"]["schema"])
            self.assertEqual(1, run["daily_summary"]["summary"]["experiment_count"])
            self.assertTrue(run["apply_audit"]["valid"], run["apply_audit"])
            self.assertEqual(8, len(run["batch_update_requests"]))
            self.assertEqual(102, run["batch_update_requests"][0]["appendCells"]["sheetId"])
            self.assertEqual(333, run["batch_update_requests"][3]["appendCells"]["sheetId"])
            self.assertEqual(101, run["batch_update_requests"][4]["updateCells"]["start"]["sheetId"])
            self.assertEqual(1, len(client.batch_updates))

    def test_live_google_daily_agent_watch_skips_duplicate_batches(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            works_path = write_fake_litscout_export(Path(tmpdir) / "works.json")
            client = FakeSheetsApiClient(
                snapshot_from_tables(
                    load_workbook_tables(workbook_path),
                    {
                        "Experiments": 101,
                        "Results": 102,
                        "Literature Evidence": 111,
                        "Agent Suggestions": 222,
                        "Daily Reviews": 333,
                    },
                )
            )

            run = run_live_google_daily_agent_watch(
                "spreadsheet-123",
                client,
                config=AgentRunConfig(review_date="2026-06-09", litscout_export=str(works_path)),
                apply=True,
                iterations=2,
                interval_seconds=0,
            )

            self.assertEqual("lab-notebook-agent-live-google-daily-watch.v1", run["schema"])
            self.assertEqual(2, run["summary"]["iteration_count"])
            self.assertEqual(1, run["summary"]["applied_iterations"])
            self.assertEqual(1, run["summary"]["duplicate_batches_skipped"])
            self.assertEqual(1, len(client.batch_updates))
            self.assertTrue(run["runs"][0]["applied"])
            self.assertEqual("duplicate_batch", run["runs"][1]["apply_skip_reason"])

    def test_live_google_daily_agent_watch_cli_writes_report(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            output = Path(tmpdir) / "watch.json"
            client = FakeSheetsApiClient(
                snapshot_from_tables(
                    load_workbook_tables(workbook_path),
                    {
                        "Experiments": 101,
                        "Results": 102,
                        "Agent Suggestions": 222,
                        "Daily Reviews": 333,
                    },
                )
            )

            with patch("lab_notebook_agent.google_api.GoogleSheetsApiClient.from_credentials", return_value=client):
                exit_code = main(
                    [
                        "google-daily-agent-watch-live",
                        "--spreadsheet-id",
                        "spreadsheet-123",
                        "--review-date",
                        "2026-06-09",
                        "--iterations",
                        "1",
                        "--interval-seconds",
                        "0",
                        "--run-output",
                        str(output),
                    ]
                )

            report = json.loads(output.read_text(encoding="utf-8"))
            self.assertEqual(0, exit_code)
            self.assertEqual("lab-notebook-agent-live-google-daily-watch.v1", report["schema"])
            self.assertEqual(1, report["summary"]["iteration_count"])

    def test_live_google_formulation_normalization_applies_valid_batch_with_fake_client(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            workbook = load_workbook(workbook_path)
            workbook["Master Reagents"]["F2"] = "156.18"
            workbook["Master Reagents"]["G2"] = "1.05"
            workbook["Formulations"]["E2"] = "10"
            workbook.save(workbook_path)
            client = FakeSheetsApiClient(
                snapshot_from_tables(
                    load_workbook_tables(workbook_path),
                    {"Formulations": 103},
                )
            )
            run = run_live_google_formulation_normalization(
                "spreadsheet-123",
                client,
                experiment_ids=("EP-001",),
                apply=True,
            )
            self.assertTrue(run["applied"])
            self.assertEqual(
                "lab-notebook-agent-formulation-normalization.v1",
                run["formulation_normalization_report"]["schema"],
            )
            self.assertEqual(2, run["formulation_normalization_report"]["summary"]["formulation_cells_to_update"])
            self.assertTrue(run["apply_audit"]["valid"], run["apply_audit"])
            self.assertEqual(2, len(run["batch_update_requests"]))
            self.assertEqual(103, run["batch_update_requests"][0]["updateCells"]["start"]["sheetId"])
            self.assertEqual(5, run["batch_update_requests"][0]["updateCells"]["start"]["columnIndex"])
            self.assertEqual(6, run["batch_update_requests"][1]["updateCells"]["start"]["columnIndex"])
            self.assertEqual(1, len(client.batch_updates))

    def test_live_google_daily_log_results_normalization_applies_valid_batch_with_fake_client(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            client = FakeSheetsApiClient(
                snapshot_from_tables(
                    load_workbook_tables(workbook_path),
                    {"Results": 102},
                )
            )
            run = run_live_google_daily_log_results_normalization(
                "spreadsheet-123",
                client,
                experiment_ids=("EP-001",),
                review_date="2026-06-09",
                apply=True,
            )
            self.assertTrue(run["applied"])
            self.assertEqual("lab-notebook-agent-daily-log-results.v1", run["daily_log_results_report"]["schema"])
            self.assertEqual(2, run["daily_log_results_report"]["summary"]["result_rows_to_append"])
            self.assertTrue(run["apply_audit"]["valid"], run["apply_audit"])
            self.assertEqual(1, len(run["batch_update_requests"]))
            self.assertEqual(102, run["batch_update_requests"][0]["appendCells"]["sheetId"])
            self.assertEqual(2, len(run["batch_update_requests"][0]["appendCells"]["rows"]))
            self.assertEqual(1, len(client.batch_updates))

    def test_live_google_material_scaffold_applies_valid_batch_with_fake_client(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            client = FakeSheetsApiClient(
                snapshot_from_tables(
                    load_workbook_tables(workbook_path),
                    {
                        "Master Reagents": 100,
                        "Formulations": 103,
                    },
                )
            )
            run = run_live_google_material_scaffold(
                "spreadsheet-123",
                client,
                experiment_id="EP-002",
                process_type="emulsion polymerization",
                apply=True,
            )
            self.assertTrue(run["applied"])
            self.assertEqual("lab-notebook-agent-live-google-material-scaffold.v1", run["schema"])
            self.assertEqual("lab-notebook-agent-material-scaffold.v1", run["material_scaffold_report"]["schema"])
            self.assertEqual(1, run["material_scaffold_report"]["summary"]["master_reagent_rows_to_append"])
            self.assertEqual(4, run["material_scaffold_report"]["summary"]["formulation_rows_to_append"])
            self.assertTrue(run["apply_audit"]["valid"], run["apply_audit"])
            self.assertEqual(2, len(run["batch_update_requests"]))
            self.assertEqual(100, run["batch_update_requests"][0]["appendCells"]["sheetId"])
            self.assertEqual(103, run["batch_update_requests"][1]["appendCells"]["sheetId"])
            self.assertEqual(1, len(client.batch_updates))

    def test_live_google_material_scaffold_cli_writes_report(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            output = Path(tmpdir) / "material-scaffold-live.json"
            client = FakeSheetsApiClient(
                snapshot_from_tables(
                    load_workbook_tables(workbook_path),
                    {
                        "Master Reagents": 100,
                        "Formulations": 103,
                    },
                )
            )

            with patch("lab_notebook_agent.google_api.GoogleSheetsApiClient.from_credentials", return_value=client):
                exit_code = main(
                    [
                        "google-scaffold-materials-live",
                        "--spreadsheet-id",
                        "spreadsheet-123",
                        "--experiment-id",
                        "EP-002",
                        "--process-type",
                        "emulsion polymerization",
                        "--apply",
                        "--run-output",
                        str(output),
                    ]
                )

            report = json.loads(output.read_text(encoding="utf-8"))
            self.assertEqual(0, exit_code)
            self.assertEqual("lab-notebook-agent-live-google-material-scaffold.v1", report["schema"])
            self.assertTrue(report["applied"])
            self.assertEqual(4, report["material_scaffold_report"]["summary"]["formulation_rows_to_append"])

    def test_live_google_plan_materialization_applies_valid_batch_with_fake_client(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            works_path = write_fake_litscout_export(Path(tmpdir) / "works.json")
            accepted_path = Path(tmpdir) / "accepted.xlsx"
            run_workbook_agent(
                workbook_path,
                AgentRunConfig(experiment_ids=("EP-001",), litscout_export=str(works_path)),
                apply=True,
                output_workbook=accepted_path,
            )
            mark_first_suggestion_accepted(accepted_path)
            client = FakeSheetsApiClient(
                snapshot_from_tables(
                    load_workbook_tables(accepted_path),
                    {
                        "Experiments": 101,
                        "Formulations": 103,
                        "Results": 102,
                        "Literature Evidence": 111,
                        "Agent Suggestions": 222,
                    },
                )
            )
            run = run_live_google_plan_materialization(
                "spreadsheet-123",
                client,
                planned_date="2026-06-10",
                apply=True,
            )
            self.assertTrue(run["applied"])
            self.assertEqual(1, run["materialization_report"]["summary"]["experiment_rows_to_append"])
            self.assertEqual(3, run["materialization_report"]["summary"]["formulation_rows_to_append"])
            self.assertEqual(8, run["materialization_report"]["summary"]["result_rows_to_append"])
            self.assertEqual(101, run["batch_update_requests"][0]["appendCells"]["sheetId"])
            self.assertEqual(103, run["batch_update_requests"][1]["appendCells"]["sheetId"])
            self.assertEqual(102, run["batch_update_requests"][2]["appendCells"]["sheetId"])

    def test_accepted_suggestion_materializes_planned_experiment_and_results(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            works_path = write_fake_litscout_export(Path(tmpdir) / "works.json")
            suggested_path = Path(tmpdir) / "suggested.xlsx"
            run_workbook_agent(
                workbook_path,
                AgentRunConfig(experiment_ids=("EP-001",), litscout_export=str(works_path)),
                apply=True,
                output_workbook=suggested_path,
            )
            mark_first_suggestion_accepted(suggested_path)
            report = build_plan_materialization_report(
                load_workbook_tables(suggested_path),
                planned_date="2026-06-10",
            )
            self.assertEqual(1, report["summary"]["experiment_rows_to_append"])
            self.assertEqual(3, report["summary"]["formulation_rows_to_append"])
            self.assertEqual(8, report["summary"]["result_rows_to_append"])
            self.assertEqual("EP-001-FUP-001", report["runs"][0]["append_experiments"][0]["experiment_id"])

            planned_path = Path(tmpdir) / "planned.xlsx"
            apply_plan_materialization_report_to_workbook(suggested_path, report, planned_path)
            workbook = load_workbook(planned_path)
            self.assertEqual("EP-001-FUP-001", workbook["Experiments"]["A3"].value)
            self.assertEqual("2026-06-10", workbook["Experiments"]["B3"].value)
            self.assertEqual("EP-001-FUP-001", workbook["Formulations"]["A5"].value)
            self.assertEqual("M-SKA", workbook["Formulations"]["B5"].value)
            self.assertEqual("225", workbook["Formulations"]["M5"].value)
            self.assertEqual("particle_size_nm", workbook["Results"]["C3"].value)
            headers = [cell.value for cell in workbook["Agent Suggestions"][1]]
            self.assertEqual(
                "run_planned",
                workbook["Agent Suggestions"].cell(row=2, column=headers.index("status") + 1).value,
            )

    def test_accepted_suggestion_materialization_derives_stock_quantities_from_master_reagents(self) -> None:
        tables = {
            "Master Reagents": [
                {
                    "reagent_id": "I-APS",
                    "concentration": "0.1",
                    "concentration_units": "M",
                }
            ],
            "Experiments": [],
            "Formulations": [],
            "Results": [],
            "Agent Suggestions": [
                {
                    "suggestion_id": "SUG-STOCK",
                    "experiment_id": "EP-001",
                    "proposed_experiment_id": "EP-001-FUP-001",
                    "status": "accepted",
                    "proposed_plan_json": json.dumps(
                        {
                            "suggested_experiment_id": "EP-001-FUP-001",
                            "process_type": "emulsion polymerization",
                            "sheet_rows": {
                                "experiments": [
                                    {
                                        "experiment_id": "EP-001-FUP-001",
                                        "process_type": "emulsion polymerization",
                                    }
                                ],
                                "formulations": [
                                    {
                                        "experiment_id": "EP-001-FUP-001",
                                        "reagent_id": "I-APS",
                                        "phase": "initiator feed",
                                        "target_role": "initiator",
                                        "volume_mL": "2.5",
                                        "moles_mmol": "",
                                    }
                                ],
                            },
                        }
                    ),
                }
            ],
        }

        report = build_plan_materialization_report(tables, planned_date="2026-06-10")

        formulation = report["runs"][0]["append_formulations"][0]
        self.assertEqual("0.1", formulation["concentration"])
        self.assertEqual("M", formulation["concentration_units"])
        self.assertEqual("0.25", formulation["moles_mmol"])
        self.assertIn("Derived missing quantity cells", formulation["notes"])

    def test_accepted_suggestion_materialization_derives_wt_percent_from_mass_total(self) -> None:
        tables = {
            "Master Reagents": [
                {
                    "reagent_id": "W-DI",
                    "density_g_mL": "1.0",
                }
            ],
            "Experiments": [],
            "Formulations": [],
            "Results": [],
            "Agent Suggestions": [
                {
                    "suggestion_id": "SUG-WT",
                    "experiment_id": "EP-001",
                    "proposed_experiment_id": "EP-001-FUP-001",
                    "status": "accepted",
                    "proposed_plan_json": json.dumps(
                        {
                            "suggested_experiment_id": "EP-001-FUP-001",
                            "process_type": "emulsion polymerization",
                            "sheet_rows": {
                                "experiments": [
                                    {
                                        "experiment_id": "EP-001-FUP-001",
                                        "process_type": "emulsion polymerization",
                                    }
                                ],
                                "formulations": [
                                    {
                                        "experiment_id": "EP-001-FUP-001",
                                        "reagent_id": "M-SKA",
                                        "phase": "monomer feed",
                                        "target_role": "core_monomer",
                                        "mass_g": "10",
                                        "wt_percent": "",
                                    },
                                    {
                                        "experiment_id": "EP-001-FUP-001",
                                        "reagent_id": "I-APS",
                                        "phase": "initiator feed",
                                        "target_role": "initiator",
                                        "mass_g": "0.2",
                                        "wt_percent": "",
                                    },
                                    {
                                        "experiment_id": "EP-001-FUP-001",
                                        "reagent_id": "W-DI",
                                        "phase": "aqueous",
                                        "target_role": "solvent",
                                        "volume_mL": "89.8",
                                        "mass_g": "",
                                        "wt_percent": "",
                                    },
                                ],
                            },
                        }
                    ),
                }
            ],
        }

        report = build_plan_materialization_report(tables, planned_date="2026-06-10")

        formulations = {
            row["reagent_id"]: row
            for row in report["runs"][0]["append_formulations"]
        }
        self.assertEqual("89.8", formulations["W-DI"]["mass_g"])
        self.assertEqual("10", formulations["M-SKA"]["wt_percent"])
        self.assertEqual("0.2", formulations["I-APS"]["wt_percent"])
        self.assertEqual("89.8", formulations["W-DI"]["wt_percent"])
        self.assertIn("Derived wt_percent", formulations["M-SKA"]["notes"])
        self.assertIn("Derived wt_percent", formulations["W-DI"]["notes"])

    def test_accepted_suggestion_materialization_emits_google_batch_requests(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = save_workbook(Path(tmpdir) / "template.xlsx")
            works_path = write_fake_litscout_export(Path(tmpdir) / "works.json")
            suggested_path = Path(tmpdir) / "suggested.xlsx"
            run_workbook_agent(
                workbook_path,
                AgentRunConfig(experiment_ids=("EP-001",), litscout_export=str(works_path)),
                apply=True,
                output_workbook=suggested_path,
            )
            mark_first_suggestion_accepted(suggested_path)
            snapshot = snapshot_from_tables(
                load_workbook_tables(suggested_path),
                {
                    "Experiments": 101,
                    "Formulations": 103,
                    "Results": 102,
                    "Literature Evidence": 111,
                    "Agent Suggestions": 222,
                },
            )
            report = build_plan_materialization_report(
                snapshot_to_tables(snapshot),
                planned_date="2026-06-10",
            )
            audit = audit_report_against_snapshot(report, snapshot)
            self.assertTrue(audit["valid"], audit["errors"])
            self.assertEqual(3, audit["summary"]["formulation_rows_to_append"])
            requests = batch_update_requests_from_report(report, sheet_ids_from_snapshot(snapshot))
            self.assertEqual(4, len(requests))
            self.assertEqual(101, requests[0]["appendCells"]["sheetId"])
            self.assertEqual(103, requests[1]["appendCells"]["sheetId"])
            self.assertEqual(102, requests[2]["appendCells"]["sheetId"])
            self.assertEqual(222, requests[3]["updateCells"]["start"]["sheetId"])


def sample_experiment_record() -> dict[str, object]:
    return {
        "experiment": {
            "experiment_id": "EP-010",
            "date": "2026-06-10",
            "project": "SABER CCSP",
            "process_type": "emulsion polymerization",
            "objective": "Record a completed small-particle latex run.",
            "hypothesis": "Higher surfactant active basis should reduce particle size.",
            "status": "complete",
            "summary": "Latex stayed stable through workup.",
        },
        "formulation": [
            {
                "reagent_id": "S-SDS",
                "phase": "aqueous",
                "target_role": "surfactant",
                "mass_g": "0.35",
                "feed_order": "0",
            }
        ],
        "observations": [
            {
                "timestamp": "2026-06-10T14:00:00",
                "stage": "feed",
                "temperature_C": "70",
                "rpm": "250",
                "observation": "Feed stayed stable; particle size 310 nm by quick DLS.",
            },
            "No visible coagulum after filtration.",
        ],
        "results": [
            {
                "measurement_type": "DLS particle size",
                "method": "intensity average",
                "value": "310",
                "units": "nm",
                "condition": "post-workup",
            }
        ],
    }


def write_fake_litscout_export(path: Path) -> Path:
    path.write_text(
        json_dumps(
            [
                {
                    "title": "Role of anionic and nonionic surfactants on particle size",
                    "service": "openalex",
                    "author_names": ["A. Author", "B. Author"],
                    "year": 2006,
                    "doi": "10.1002/app.23717",
                    "url": "https://doi.org/10.1002/app.23717",
                    "cited_by_count": 10,
                    "concepts": [
                        {"display_name": "Emulsion polymerization"},
                        {"display_name": "Particle size"},
                        {"display_name": "Surfactant"},
                    ],
                }
            ]
        ),
        encoding="utf-8",
    )
    return path


def mark_first_suggestion_accepted(path: Path) -> None:
    workbook = load_workbook(path)
    worksheet = workbook["Agent Suggestions"]
    headers = [cell.value for cell in worksheet[1]]
    status_column = headers.index("status") + 1
    worksheet.cell(row=2, column=status_column, value="accepted")
    workbook.save(path)


def set_agent_config(tables: dict[str, list[dict[str, object]]], key: str, value: str) -> None:
    rows = tables.setdefault("Agent Config", [])
    for row in rows:
        if str(row.get("key", "")).strip() == key:
            row["value"] = value
            return
    rows.append({"key": key, "value": value, "notes": ""})


def complete_template_materials(tables: dict[str, list[dict[str, object]]]) -> None:
    for row in tables["Master Reagents"]:
        if row["reagent_id"] == "M-SKA":
            row["molecular_weight_g_mol"] = "156.18"
            row["density_g_mL"] = "1.05"
        if row["reagent_id"] == "S-SDS":
            row["concentration"] = "active mass"
    for row in tables["Formulations"]:
        if row["reagent_id"] == "M-SKA":
            row["mass_g"] = "10"
        if row["reagent_id"] == "I-APS":
            row["mass_g"] = "0.2"
        if row["reagent_id"] == "S-SDS":
            row["mass_g"] = "0.1"


def low_confidence_agent_tables(confidence_floor: str) -> dict[str, list[dict[str, object]]]:
    return {
        "Experiments": [
            {
                "experiment_id": "GEN-001",
                "date": "2026-06-09",
                "process_type": "miscellaneous screening",
                "objective": "Capture an exploratory note without process-specific outcome signals.",
                "status": "complete",
            }
        ],
        "Master Reagents": [],
        "Formulations": [],
        "Daily Log": [],
        "Results": [],
        "Literature Evidence": [],
        "Agent Suggestions": [],
        "Agent Config": [
            {"key": "suggestion_confidence_floor", "value": confidence_floor, "notes": ""},
        ],
    }


def write_tables_to_workbook(path: Path, tables: dict[str, list[dict[str, object]]]) -> Path:
    workbook_path = save_workbook(path, include_examples=False)
    workbook = load_workbook(workbook_path)
    for sheet_name, rows in tables.items():
        worksheet = workbook[sheet_name]
        headers = [cell.value for cell in worksheet[1]]
        for row in rows:
            worksheet.append([row.get(header, "") for header in headers])
    workbook.save(workbook_path)
    return workbook_path


def validations_by_range(worksheet: object) -> dict[str, str]:
    return {
        str(validation.sqref): str(validation.formula1)
        for validation in worksheet.data_validations.dataValidation
    }


class FakeSheetsApiClient:
    def __init__(self, snapshot: dict[str, object]) -> None:
        self.snapshot = snapshot
        self.batch_updates: list[list[dict[str, object]]] = []

    def get_metadata(self, spreadsheet_id: str) -> dict[str, object]:
        sheets = []
        for sheet_name, payload in self.snapshot["sheets"].items():
            sheets.append(
                {
                    "properties": {
                        "title": sheet_name,
                        "sheetId": payload.get("sheet_id"),
                    }
                }
            )
        return {
            "spreadsheetId": spreadsheet_id,
            "sheets": sheets,
        }

    def get_values(
        self,
        spreadsheet_id: str,
        sheet_name: str,
        value_range: str,
        value_render_option: str = "FORMATTED_VALUE",
    ) -> list[list[object]]:
        return self.snapshot["sheets"][sheet_name]["values"]

    def batch_update(self, spreadsheet_id: str, requests: list[dict[str, object]]) -> dict[str, object]:
        self.batch_updates.append(requests)
        return {
            "spreadsheetId": spreadsheet_id,
            "replies": [{} for _ in requests],
        }


def json_dumps(value: object) -> str:
    import json

    return json.dumps(value, indent=2) + "\n"


if __name__ == "__main__":
    unittest.main()
